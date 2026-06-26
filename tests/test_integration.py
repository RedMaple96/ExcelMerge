"""集成测试（端到端）— 覆盖验收标准 AC-01 ~ AC-11。

测试目标：以“创建 xlsx 夹具 -> ExcelLoader 加载 -> ExcelComparator 比较 ->
ExcelMerger 合并 -> 保存 -> 重载 -> 校验”的完整管线方式，端到端验证核心
能力。不测试 GUI（GUI 需要图形显示环境），仅测试 GUI 所调用的核心管线。

AC 映射：
- AC-01 样式保留：合并后值变更但样式保持。
- AC-02 公式存活：合并后公式字符串原样保留，不被计算值替代。
- AC-03 多 Sheet：右侧新增/修改工作表能合并到左侧。
- AC-04 追加行：右侧新增行按 key 对齐后追加到左侧。
- AC-05 同步滚动：GUI 功能，跳过。
- AC-06 差异高亮：DiffResult 正确识别 different/left_only/right_only 及 diff_cells。
- AC-07 差异导航：diff_row_indices 非空且有序。
- AC-08 拖拽打开：GUI 功能，跳过。
- AC-09 大文件比较：10000x50 比较耗时 < 30s。
- AC-10 界面响应：GUI 功能，跳过。
- AC-11 内存：10000x50 比较峰值内存 < 500MB。

补充端到端用例：完整往返、忽略列、关键列对齐、备份创建、保存后完整性。
"""

import os
import shutil
import tempfile
import time
import tracemalloc
import unittest
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from src.core.comparator import ExcelComparator
from src.core.excel_loader import ExcelLoader
from src.core.merger import ExcelMerger


# GUI 测试需要图形显示环境；本集成测试只覆盖核心管线，不运行 GUI。
GUI_AVAILABLE = False


class IntegrationTest(unittest.TestCase):
    """端到端集成测试。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup)

    def _cleanup(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _path(self, name: str) -> str:
        return os.path.join(self.tmpdir, name)

    # ------------------------------------------------------------------
    # 夹具构建辅助函数
    # ------------------------------------------------------------------
    @staticmethod
    def _style_cell(cell, value, fill_rgb="FFFF0000", bold=True, border_style="thin"):
        """给单元格赋值并设置：红色实心填充 + 加粗字体 + 四边细边框。"""
        cell.value = value
        cell.fill = PatternFill(
            start_color=fill_rgb, end_color=fill_rgb, fill_type="solid"
        )
        cell.font = Font(bold=bold)
        side = Side(style=border_style)
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def _write_large_workbook(path, rows, cols, diff_rows=None):
        """用 write_only 模式快速生成大工作簿。

        - diff_rows: 需要在第 1 列制造差异的行号集合(1-based)；
          为空则所有单元格使用基准值 "r{行}c{列}"。
        """
        diff_rows = diff_rows or set()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")
        for r in range(1, rows + 1):
            row_vals = []
            for c in range(1, cols + 1):
                if c == 1 and r in diff_rows:
                    row_vals.append("DIFF_{}".format(r))
                else:
                    row_vals.append("r{}c{}".format(r, c))
            ws.append(row_vals)
        wb.save(path)

    @staticmethod
    def _copy_whole_sheet(src_ws, tgt_ws):
        """把源工作表所有单元格的值与样式复制到目标工作表（跨工作簿）。"""
        for r in range(1, src_ws.max_row + 1):
            for c in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row=r, column=c)
                tgt_cell = tgt_ws.cell(row=r, column=c, value=src_cell.value)
                ExcelLoader.copy_cell_style(src_cell, tgt_cell)

    # ------------------------------------------------------------------
    # AC-01 样式保留：合并后值变更，样式（填充/字体/边框）保持
    # ------------------------------------------------------------------
    def test_ac01_style_preservation(self):
        left_path = self._path("ac01_left.xlsx")
        right_path = self._path("ac01_right.xlsx")

        wb_l = Workbook()
        self._style_cell(wb_l.active["A1"], "left_value")
        wb_l.save(left_path)

        wb_r = Workbook()
        # 右侧值不同但样式与左侧一致
        self._style_cell(wb_r.active["A1"], "right_value")
        wb_r.save(right_path)

        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_r2 = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right)
        self.assertEqual(diff.aligned_rows[0].status, "different")

        ExcelMerger.merge_right_to_left(diff, left, right)
        ExcelLoader.save_workbook(wb_l2, left_path)

        wb_final = ExcelLoader.load_workbook(left_path)
        cell = wb_final.active["A1"]
        # 值已更新为右侧
        self.assertEqual(cell.value, "right_value")
        # 样式保留：红色填充、加粗、四边细边框
        self.assertIn("FF0000", cell.fill.fgColor.rgb)
        self.assertTrue(cell.font.bold)
        self.assertEqual(cell.border.left.style, "thin")
        self.assertEqual(cell.border.right.style, "thin")
        self.assertEqual(cell.border.top.style, "thin")
        self.assertEqual(cell.border.bottom.style, "thin")

    # ------------------------------------------------------------------
    # AC-02 公式存活：合并后公式字符串原样保留，非计算值
    # ------------------------------------------------------------------
    def test_ac02_formula_survival(self):
        left_path = self._path("ac02_left.xlsx")
        right_path = self._path("ac02_right.xlsx")

        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        ws_l["B1"] = 2
        ws_l["C1"] = "=A1+B1"
        wb_l.save(left_path)

        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 1
        ws_r["B1"] = 2
        ws_r["C1"] = "=A2+B2"
        wb_r.save(right_path)

        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_r2 = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right)
        # 仅 C 列(索引 2)不同
        self.assertEqual(diff.aligned_rows[0].diff_cells, [2])

        ExcelMerger.merge_right_to_left(diff, left, right)
        ExcelLoader.save_workbook(wb_l2, left_path)

        # data_only=False 保留公式字符串
        wb_final = ExcelLoader.load_workbook(left_path)
        self.assertEqual(wb_final.active["C1"].value, "=A2+B2")

    # ------------------------------------------------------------------
    # AC-03 多 Sheet：右侧修改 Sheet1、新增 Sheet3，合并后左侧含三个表
    # ------------------------------------------------------------------
    def test_ac03_multi_sheet_merge(self):
        left_path = self._path("ac03_left.xlsx")
        right_path = self._path("ac03_right.xlsx")

        wb_l = Workbook()
        wb_l.active.title = "Sheet1"
        wb_l["Sheet1"]["A1"] = "old"
        ws_l2 = wb_l.create_sheet("Sheet2")
        ws_l2["A1"] = "keep_sheet2"
        wb_l.save(left_path)

        wb_r = Workbook()
        wb_r.active.title = "Sheet1"
        wb_r["Sheet1"]["A1"] = "new"
        ws_r3 = wb_r.create_sheet("Sheet3")
        ws_r3["A1"] = "new_sheet3"
        wb_r.save(right_path)

        wb_l = ExcelLoader.load_workbook(left_path)
        wb_r = ExcelLoader.load_workbook(right_path)

        # 编排（镜像 GUI 按表处理的逻辑）：
        # 遍历右侧每个工作表，存在则比较+右覆盖，不存在则新建并整表复制
        for sheet_name in ExcelLoader.get_sheet_names(wb_r):
            if sheet_name in ExcelLoader.get_sheet_names(wb_l):
                left_sd = ExcelLoader.extract_sheet_data(
                    ExcelLoader.get_worksheet(wb_l, sheet_name)
                )
                right_sd = ExcelLoader.extract_sheet_data(
                    ExcelLoader.get_worksheet(wb_r, sheet_name)
                )
                diff = ExcelComparator.compare_sheets(left_sd, right_sd)
                ExcelMerger.merge_right_to_left(diff, left_sd, right_sd)
            else:
                new_ws = wb_l.create_sheet(sheet_name)
                self._copy_whole_sheet(
                    ExcelLoader.get_worksheet(wb_r, sheet_name), new_ws
                )

        ExcelLoader.save_workbook(wb_l, left_path)

        wb_final = ExcelLoader.load_workbook(left_path)
        # 三个工作表，名称与顺序正确
        self.assertEqual(
            ExcelLoader.get_sheet_names(wb_final), ["Sheet1", "Sheet2", "Sheet3"]
        )
        self.assertEqual(wb_final["Sheet1"]["A1"].value, "new")
        self.assertEqual(wb_final["Sheet2"]["A1"].value, "keep_sheet2")
        self.assertEqual(wb_final["Sheet3"]["A1"].value, "new_sheet3")

    # ------------------------------------------------------------------
    # AC-04 追加行：右侧新增 ID=4，按 key 列对齐后追加到左侧
    # ------------------------------------------------------------------
    def test_ac04_append_rows(self):
        left_path = self._path("ac04_left.xlsx")
        right_path = self._path("ac04_right.xlsx")

        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        ws_l["A2"] = 2
        ws_l["A3"] = 3
        wb_l.save(left_path)

        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 1
        ws_r["A2"] = 2
        ws_r["A3"] = 3
        ws_r["A4"] = 4
        ws_r["A4"].font = Font(bold=True)
        ws_r["A4"].fill = PatternFill(
            start_color="FF0000FF", end_color="FF0000FF", fill_type="solid"
        )
        wb_r.save(right_path)

        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_r2 = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right, key_cols=[0])
        self.assertEqual(diff.stats["right_only"], 1)

        ExcelMerger.append_rows(diff, left, right)
        ExcelLoader.save_workbook(wb_l2, left_path)

        wb_final = ExcelLoader.load_workbook(left_path)
        ws = wb_final.active
        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws["A4"].value, 4)
        # 样式应来自右侧第 4 行
        self.assertTrue(ws["A4"].font.bold)
        self.assertIn("0000FF", ws["A4"].fill.fgColor.rgb)

    # ------------------------------------------------------------------
    # AC-05 同步滚动：GUI 渲染功能，跳过
    # ------------------------------------------------------------------
    @unittest.skipUnless(GUI_AVAILABLE, "AC-05 同步滚动为 GUI 功能，需图形显示环境")
    def test_ac05_sync_scroll(self):
        # GUI 测试不在核心管线范围内
        self.skipTest("GUI 渲染测试")

    # ------------------------------------------------------------------
    # AC-06 差异高亮：DiffResult 正确分类行并填充 diff_cells，供 GUI 高亮
    # ------------------------------------------------------------------
    def test_ac06_diff_highlight_data(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "k1"
        ws_l["B1"] = "a"  # same
        ws_l["A2"] = "k2"
        ws_l["B2"] = "b"  # different
        ws_l["A3"] = "k3"  # left_only
        wb_l.save(self._path("ac06_left.xlsx"))

        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "k1"
        ws_r["B1"] = "a"
        ws_r["A2"] = "k2"
        ws_r["B2"] = "CHANGED"
        ws_r["A3"] = "k4"  # right_only
        wb_r.save(self._path("ac06_right.xlsx"))

        wb_l2 = ExcelLoader.load_workbook(self._path("ac06_left.xlsx"))
        wb_r2 = ExcelLoader.load_workbook(self._path("ac06_right.xlsx"))
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # 统计：1 same / 1 different / 1 left_only / 1 right_only
        self.assertEqual(diff.stats["same"], 1)
        self.assertEqual(diff.stats["different"], 1)
        self.assertEqual(diff.stats["left_only"], 1)
        self.assertEqual(diff.stats["right_only"], 1)

        # different 行的 diff_cells 应被填充，GUI 据此高亮
        diff_pairs = [p for p in diff.aligned_rows if p.status == "different"]
        self.assertEqual(len(diff_pairs), 1)
        self.assertEqual(diff_pairs[0].diff_cells, [1])  # B 列(索引 1)

    # ------------------------------------------------------------------
    # AC-07 差异导航：diff_row_indices 非空且升序，供 GUI 逐项导航
    # ------------------------------------------------------------------
    def test_ac07_diff_navigation_indices(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        for i, v in enumerate(["a", "b", "c", "d"], start=1):
            ws_l.cell(row=i, column=1, value=v)
        wb_l.save(self._path("ac07_left.xlsx"))

        wb_r = Workbook()
        ws_r = wb_r.active
        # 第 1、3 行相同；第 2、4 行不同
        for i, v in enumerate(["a", "X", "c", "Y"], start=1):
            ws_r.cell(row=i, column=1, value=v)
        wb_r.save(self._path("ac07_right.xlsx"))

        wb_l2 = ExcelLoader.load_workbook(self._path("ac07_left.xlsx"))
        wb_r2 = ExcelLoader.load_workbook(self._path("ac07_right.xlsx"))
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right)

        self.assertGreater(len(diff.diff_row_indices), 0)
        # 升序
        self.assertEqual(diff.diff_row_indices, sorted(diff.diff_row_indices))
        # 第 2、4 行(0-based 1、3)为差异行
        self.assertEqual(diff.diff_row_indices, [1, 3])

    # ------------------------------------------------------------------
    # AC-08 拖拽打开：GUI 功能，跳过
    # ------------------------------------------------------------------
    @unittest.skipUnless(GUI_AVAILABLE, "AC-08 拖拽打开为 GUI 功能，需图形显示环境")
    def test_ac08_drag_open(self):
        self.skipTest("GUI 交互测试")

    # ------------------------------------------------------------------
    # AC-09 大文件比较：10000x50 比较耗时 < 30s（仅计时 compare_sheets）
    # ------------------------------------------------------------------
    def test_ac09_large_file_compare_under_30s(self):
        rows, cols = 10000, 50
        left_path = self._path("ac09_left.xlsx")
        right_path = self._path("ac09_right.xlsx")
        diff_rows = set(range(100, 200))  # 100 行在第 1 列制造差异
        self._write_large_workbook(left_path, rows, cols)
        self._write_large_workbook(right_path, rows, cols, diff_rows=diff_rows)

        wb_l = ExcelLoader.load_workbook(left_path)
        wb_r = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l.worksheets[0])
        right = ExcelLoader.extract_sheet_data(wb_r.worksheets[0])

        # 仅计时比较步骤
        start = time.time()
        diff = ExcelComparator.compare_sheets(left, right)
        elapsed = time.time() - start

        self.assertLess(elapsed, 30, "比较耗时 {:.2f}s 超过 30s".format(elapsed))
        # 应识别出 100 个差异行
        self.assertEqual(diff.stats["different"], 100)

    # ------------------------------------------------------------------
    # AC-10 界面响应：GUI 运行时特性，跳过
    # ------------------------------------------------------------------
    @unittest.skipUnless(GUI_AVAILABLE, "AC-10 界面响应为 GUI 运行时特性，需图形显示环境")
    def test_ac10_ui_responsiveness(self):
        self.skipTest("GUI 运行时测试")

    # ------------------------------------------------------------------
    # AC-11 内存：10000x50 比较峰值内存 < 500MB
    # ------------------------------------------------------------------
    def test_ac11_memory_under_500mb(self):
        rows, cols = 10000, 50
        left_path = self._path("ac11_left.xlsx")
        right_path = self._path("ac11_right.xlsx")
        diff_rows = set(range(100, 200))
        self._write_large_workbook(left_path, rows, cols)
        self._write_large_workbook(right_path, rows, cols, diff_rows=diff_rows)

        tracemalloc.start()
        try:
            wb_l = ExcelLoader.load_workbook(left_path)
            wb_r = ExcelLoader.load_workbook(right_path)
            left = ExcelLoader.extract_sheet_data(wb_l.worksheets[0])
            right = ExcelLoader.extract_sheet_data(wb_r.worksheets[0])
            ExcelComparator.compare_sheets(left, right)
            _current, peak = tracemalloc.get_traced_memory()
        finally:
            tracemalloc.stop()

        self.assertLess(
            peak,
            500 * 1024 * 1024,
            "峰值内存 {:.1f}MB 超过 500MB".format(peak / 1024 / 1024),
        )

    # ------------------------------------------------------------------
    # 补充：完整往返 — 复杂夹具(样式+公式+合并单元格+多种差异)右覆盖后全部消除
    # ------------------------------------------------------------------
    def test_full_round_trip(self):
        left_path = self._path("rt_left.xlsx")
        right_path = self._path("rt_right.xlsx")

        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "id"
        ws_l["B1"] = "name"
        ws_l["A2"] = 1
        ws_l["B2"] = "alice"  # 将不同
        ws_l["C2"] = "=A2+10"  # 公式将不同
        ws_l["A3"] = 2
        ws_l["B3"] = "bob"
        ws_l.merge_cells("A3:B3")  # 合并单元格
        ws_l["B2"].fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )
        wb_l.save(left_path)

        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "id"
        ws_r["B1"] = "name"
        ws_r["A2"] = 1
        ws_r["B2"] = "ALICE"  # different
        ws_r["C2"] = "=A2+20"  # 公式 different
        ws_r["A3"] = 2
        ws_r["B3"] = "bob"
        ws_r.merge_cells("A3:B3")
        ws_r["B2"].fill = PatternFill(
            start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"
        )
        wb_r.save(right_path)

        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_r2 = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right)
        self.assertGreater(diff.stats["different"], 0)

        ExcelMerger.merge_right_to_left(diff, left, right)
        ExcelLoader.save_workbook(wb_l2, left_path)

        # 重载后再次比较，应全部 same
        wb_l3 = ExcelLoader.load_workbook(left_path)
        wb_r3 = ExcelLoader.load_workbook(right_path)
        left2 = ExcelLoader.extract_sheet_data(wb_l3.active)
        right2 = ExcelLoader.extract_sheet_data(wb_r3.active)
        diff2 = ExcelComparator.compare_sheets(left2, right2)
        self.assertEqual(diff2.stats["same"], len(diff2.aligned_rows))
        self.assertEqual(diff2.stats["different"], 0)
        self.assertEqual(diff2.stats["left_only"], 0)
        self.assertEqual(diff2.stats["right_only"], 0)

    # ------------------------------------------------------------------
    # 补充：忽略列 — 被忽略列的差异不导致行判定为 different
    # ------------------------------------------------------------------
    def test_ignore_columns(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "x"
        ws_l["B1"] = "b1"
        ws_l["C1"] = "c1"
        wb_l.save(self._path("ig_left.xlsx"))

        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "x"
        ws_r["B1"] = "B1_CHANGED"  # B 列不同但被忽略
        ws_r["C1"] = "c1"
        wb_r.save(self._path("ig_right.xlsx"))

        wb_l2 = ExcelLoader.load_workbook(self._path("ig_left.xlsx"))
        wb_r2 = ExcelLoader.load_workbook(self._path("ig_right.xlsx"))
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right, ignore_cols=[1])
        self.assertEqual(diff.aligned_rows[0].status, "same")

    # ------------------------------------------------------------------
    # 补充：关键列对齐 — 行顺序不同但按 key 匹配后全部 same
    # ------------------------------------------------------------------
    def test_key_column_alignment(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        ws_l["B1"] = "a"
        ws_l["A2"] = 2
        ws_l["B2"] = "b"
        ws_l["A3"] = 3
        ws_l["B3"] = "c"
        wb_l.save(self._path("key_left.xlsx"))

        wb_r = Workbook()
        ws_r = wb_r.active
        # 逆序排列，内容一致
        ws_r["A1"] = 3
        ws_r["B1"] = "c"
        ws_r["A2"] = 2
        ws_r["B2"] = "b"
        ws_r["A3"] = 1
        ws_r["B3"] = "a"
        wb_r.save(self._path("key_right.xlsx"))

        wb_l2 = ExcelLoader.load_workbook(self._path("key_left.xlsx"))
        wb_r2 = ExcelLoader.load_workbook(self._path("key_right.xlsx"))
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right, key_cols=[0])
        self.assertTrue(all(p.status == "same" for p in diff.aligned_rows))
        self.assertEqual(diff.stats["same"], 3)

    # ------------------------------------------------------------------
    # 补充：备份创建 — 合并前用 shutil.copy2 备份原文件，备份内容等于原始
    # ------------------------------------------------------------------
    def test_backup_creation(self):
        left_path = self._path("bk_left.xlsx")
        wb_l = Workbook()
        wb_l.active["A1"] = "original"
        wb_l.save(left_path)

        # 读取原始文件字节用于比对
        with open(left_path, "rb") as f:
            original_bytes = f.read()

        # 加载并修改（在内存中）
        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_l2.active["A1"] = "modified"

        # 保存前备份：copy2 复制当前磁盘文件（仍为 original）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self._path("left_backup_{}.xlsx".format(timestamp))
        shutil.copy2(left_path, backup_path)

        # 保存修改
        ExcelLoader.save_workbook(wb_l2, left_path)

        # 备份存在且内容等于原始
        self.assertTrue(os.path.exists(backup_path))
        with open(backup_path, "rb") as f:
            backup_bytes = f.read()
        self.assertEqual(backup_bytes, original_bytes)

        # 当前文件已变为修改后内容
        wb_check = ExcelLoader.load_workbook(left_path)
        self.assertEqual(wb_check.active["A1"].value, "modified")

    # ------------------------------------------------------------------
    # 补充：保存后完整性 — 保存重载不抛异常，值与样式保持
    # ------------------------------------------------------------------
    def test_integrity_after_save(self):
        path = self._path("integrity.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "value"
        ws["A1"].fill = PatternFill(
            start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"
        )
        ws["A1"].font = Font(bold=True, italic=True)
        ws["B1"] = "=A1+1"
        ExcelLoader.save_workbook(wb, path)

        # 重载不抛异常
        wb2 = ExcelLoader.load_workbook(path)
        ws2 = wb2.active
        self.assertEqual(ws2["A1"].value, "value")
        self.assertIn("00FF00", ws2["A1"].fill.fgColor.rgb)
        self.assertTrue(ws2["A1"].font.bold)
        self.assertTrue(ws2["A1"].font.italic)
        self.assertEqual(ws2["B1"].value, "=A1+1")


if __name__ == "__main__":
    unittest.main()
