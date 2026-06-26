"""ExcelLoader 单元测试 — FR-01。

使用 unittest，测试数据通过 openpyxl 在临时目录中动态生成 .xlsx 文件。
覆盖：公式保留、样式加载、合并单元格填充、多工作表、保存持久化、
.xls 拒绝、样式复制、空表保护，以及字符串化与列标签等补充用例。
"""

import os
import shutil
import tempfile
import unittest
from unittest.mock import MagicMock

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.core.excel_loader import ExcelLoader, SheetData


class ExcelLoaderTest(unittest.TestCase):
    """ExcelLoader 功能测试。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup)

    def _cleanup(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _path(self, name: str) -> str:
        return os.path.join(self.tmpdir, name)

    # 1. 公式保留：data_only=False 时公式字符串应保留在 values 中
    def test_formula_preserved(self):
        path = self._path("formula.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["B1"] = 2
        ws["C1"] = "=A1+B1"
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        data = ExcelLoader.extract_sheet_data(wb2.active)
        # C1 位于第 1 行第 3 列 -> values[0][2]
        self.assertEqual(data.values[0][0], "1")
        self.assertEqual(data.values[0][1], "2")
        self.assertEqual(data.values[0][2], "=A1+B1")

    # 2. 样式加载：保存后重载，样式应可访问且不崩溃
    def test_styles_accessible_after_reload(self):
        path = self._path("styles.xlsx")
        wb = Workbook()
        cell = wb.active["A1"]
        cell.value = "styled"
        cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        cell.font = Font(bold=True)
        side = Side(style="thin")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
        cell.alignment = Alignment(horizontal="center")
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        cell2 = wb2.active["A1"]
        self.assertTrue(cell2.font.bold)
        self.assertIn("FF0000", cell2.fill.fgColor.rgb)
        self.assertEqual(cell2.alignment.horizontal, "center")
        self.assertEqual(cell2.border.left.style, "thin")

    # 3. 合并单元格：A1:B2 区域内所有单元格都应填充 A1 的值
    def test_merged_cells_filled_with_top_left_value(self):
        path = self._path("merged.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "merged_value"
        ws.merge_cells("A1:B2")
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        data = ExcelLoader.extract_sheet_data(wb2.active)
        # values[0][0]=A1, [0][1]=B1, [1][0]=A2, [1][1]=B2
        self.assertEqual(data.values[0][0], "merged_value")
        self.assertEqual(data.values[0][1], "merged_value")
        self.assertEqual(data.values[1][0], "merged_value")
        self.assertEqual(data.values[1][1], "merged_value")
        self.assertIn((1, 1, 2, 2), data.merged_ranges)

    # 4. 多工作表：get_sheet_names 返回全部工作表名
    def test_multi_sheet_names(self):
        path = self._path("multi.xlsx")
        wb = Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "hello"
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        self.assertEqual(ExcelLoader.get_sheet_names(wb2), ["Sheet1", "Sheet2"])
        ws = ExcelLoader.get_worksheet(wb2, "Sheet2")
        self.assertEqual(ws["A1"].value, "hello")

    # 5. 保存持久化：修改单元格后保存重载，变更应保留
    def test_save_persists_change(self):
        path = self._path("save.xlsx")
        wb = Workbook()
        wb.active["A1"] = "before"
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        wb2.active["A1"] = "after"
        ExcelLoader.save_workbook(wb2, path)

        wb3 = ExcelLoader.load_workbook(path)
        self.assertEqual(wb3.active["A1"].value, "after")

    # 6. .xls 拒绝：非 .xlsx 扩展名应抛 ValueError
    def test_xls_extension_rejected(self):
        with self.assertRaises(ValueError):
            ExcelLoader.load_workbook("fake.xls")
        # 无扩展名同样应被拒绝
        with self.assertRaises(ValueError):
            ExcelLoader.load_workbook("no_extension")

    # 7. copy_cell_style：复制后目标应具备相同样式，且为深拷贝
    def test_copy_cell_style_deep(self):
        wb = Workbook()
        ws = wb.active
        src = ws["A1"]
        src.value = "src"
        src.fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        src.font = Font(bold=True, italic=True, size=14)
        src.number_format = "0.00"
        tgt = ws["B1"]

        ExcelLoader.copy_cell_style(src, tgt)

        self.assertTrue(tgt.font.bold)
        self.assertTrue(tgt.font.italic)
        self.assertEqual(tgt.font.size, 14)
        self.assertIn("00FF00", tgt.fill.fgColor.rgb)
        self.assertEqual(tgt.number_format, "0.00")

        # 深拷贝验证：修改目标不应影响源
        tgt.font = Font(bold=False)
        self.assertTrue(src.font.bold)

    # 8. 空表保护：max_row=0 时不应崩溃，返回空结构
    def test_empty_sheet_does_not_crash(self):
        ws_mock = MagicMock()
        ws_mock.max_row = 0
        ws_mock.max_column = 0
        ws_mock.merged_cells.ranges = []

        data = ExcelLoader.extract_sheet_data(ws_mock)

        self.assertIsInstance(data, SheetData)
        self.assertEqual(data.max_row, 0)
        self.assertEqual(data.max_col, 0)
        self.assertEqual(data.values, [])
        self.assertEqual(data.header_labels, [])
        self.assertEqual(data.merged_ranges, [])

    # 补充：header_labels 正确生成
    def test_header_labels(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 1
        ws["C1"] = 3  # 写入 C1 使 max_column=3
        data = ExcelLoader.extract_sheet_data(ws)
        self.assertEqual(data.header_labels, ["A", "B", "C"])
        # 中间空单元格 B1 应为 ""
        self.assertEqual(data.values[0][1], "")

    # 补充：bool / None 字符串化规则
    def test_stringify_bool_and_none(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = True
        ws["B1"] = False
        ws["C1"] = None
        data = ExcelLoader.extract_sheet_data(ws)
        self.assertEqual(data.values[0][0], "TRUE")
        self.assertEqual(data.values[0][1], "FALSE")
        self.assertEqual(data.values[0][2], "")

    # 补充：copy_cell_value 公式与普通值分支
    def test_copy_cell_value_formula_and_plain(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=A2+1"
        ws["A2"] = 42
        ws["A3"] = None

        ExcelLoader.copy_cell_value(ws["A1"], ws["B1"])
        ExcelLoader.copy_cell_value(ws["A2"], ws["B2"])
        ExcelLoader.copy_cell_value(ws["A3"], ws["B3"])

        self.assertEqual(ws["B1"].value, "=A2+1")
        self.assertEqual(ws["B2"].value, 42)
        self.assertIsNone(ws["B3"].value)


if __name__ == "__main__":
    unittest.main()
