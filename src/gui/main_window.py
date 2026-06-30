"""主窗口模块 — MainWindow(QMainWindow)。

应用外壳：菜单栏 / 工具栏 / 状态栏 / 左右并排双表格视图。
负责文件加载、Sheet 切换、比较触发与同步滚动；差异可视化（Task 6）、
合并交互（Task 8）、保存备份（Task 9）等在此预留方法桩。
对应需求：FR-03（主窗口框架与双表格视图）。
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime
from typing import List, Optional, Tuple

from PySide6.QtCore import Qt
from PySide6.QtGui import (
    QAction,
    QBrush,
    QColor,
    QCursor,
    QGuiApplication,
    QKeySequence,
    QShortcut,
)
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMenu,
    QMessageBox,
    QProgressBar,
    QSizePolicy,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QToolBar,
    QVBoxLayout,
    QWidget,
)
from openpyxl.utils import get_column_letter

from src.core.comparator import ColPair, DiffResult, ExcelComparator
from src.core.excel_loader import ExcelLoader, SheetData
from src.core.merger import ExcelMerger
from src.gui.bottom_bar import BottomBar
from src.gui.column_settings_dialog import ColumnSettingsDialog
from src.gui.diff_birds_eye import DiffBirdsEyeView
from src.gui.diff_col_birds_eye import DiffColBirdsEyeView
from src.gui.themes import is_dark_mode
from src.gui.workers import CompareWorker


class MainWindow(QMainWindow):
    """应用程序主窗口。"""

    # ------------------------------------------------------------------ #
    # 生命周期
    # ------------------------------------------------------------------ #
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("表格比较 — 无标题")
        self.resize(1200, 750)

        # ---- 共享状态 ----
        self.left_path: Optional[str] = None
        self.right_path: Optional[str] = None
        self.left_wb = None
        self.right_wb = None
        self.left_sheet_data: Optional[SheetData] = None
        self.right_sheet_data: Optional[SheetData] = None
        self.left_sheet_name: str = ""
        self.right_sheet_name: str = ""
        self.diff_result: Optional[DiffResult] = None
        self.key_cols: List[int] = []
        self.ignore_cols: List[int] = []
        self.show_row_numbers: bool = True
        self._sync_scroll_blocked: bool = False  # 同步滚动防递归守卫
        self._current_diff_pos: int = -1  # 差异导航游标（指向 diff_row_indices）
        self._compare_worker: Optional[CompareWorker] = None  # 比较后台线程
        self._dirty: bool = False  # 是否有未保存的修改（合并后置 True）
        self._backup_done: bool = True  # 当前脏状态是否已备份；初始无需备份
        self._last_backup_path: Optional[str] = None  # 最近一次备份路径
        # 右键菜单上下文（由 on_context_menu 填充，_ctx_* 回调读取）
        self._ctx_table: Optional[QTableWidget] = None
        self._ctx_aligned_idx: int = -1  # 对齐行索引（aligned_rows 下标）
        self._ctx_row: int = -1  # 视觉行
        self._ctx_col: int = -1  # 视觉列

        # ---- 底部导航栏相关状态 ----
        self._sheet_diff_summary: dict = {}  # sheet_name -> has_diff (bool)

        # ---- 构建界面 ----
        self._init_ui()
        self._init_menubar()
        self._init_toolbar()
        self._init_statusbar()
        self._init_connections()
        self._init_shortcuts()
        self._init_theme_listener()

    # ------------------------------------------------------------------ #
    # 中央界面（左右双面板）
    # ------------------------------------------------------------------ #
    def _init_ui(self) -> None:
        """构建中央控件：左右两个面板 + 底部导航栏。"""
        central = QWidget(self)
        self.setCentralWidget(central)
        central.setAcceptDrops(True)  # 启用窗口级拖拽接收

        outer = QVBoxLayout(central)
        outer.setContentsMargins(4, 4, 4, 0)
        outer.setSpacing(0)

        # ---- 上方：左右双面板 ----
        panels = QHBoxLayout()
        panels.setContentsMargins(0, 0, 0, 0)
        panels.setSpacing(4)

        # 左侧面板
        (
            self.left_file_label,
            self.left_sheet_combo,
            self.left_table,
        ) = self._make_panel(panels, "左侧")

        # 右侧面板
        (
            self.right_file_label,
            self.right_sheet_combo,
            self.right_table,
        ) = self._make_panel(panels, "右侧")

        # 最左侧：全局差异缩略图导航条（极窄纵列，红色刻度标记差异行）
        self.diff_birds_eye = DiffBirdsEyeView(self.left_table)
        panels.insertWidget(0, self.diff_birds_eye)

        outer.addLayout(panels, 1)

        # ---- 下方：列方向差异缩略图导航条（极窄横条，标记差异列）----
        self.diff_col_birds_eye = DiffColBirdsEyeView(self.left_table)
        outer.addWidget(self.diff_col_birds_eye)

        # ---- 下方：底部导航栏 ----
        self.bottom_bar = BottomBar(self)
        outer.addWidget(self.bottom_bar)

        self._configure_table(self.left_table)
        self._configure_table(self.right_table)

    def _make_panel(self, parent_layout: QHBoxLayout, title: str):
        """创建单个面板，返回 (file_label, sheet_combo, table)。"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(4)

        # 标题区：文件信息 + Sheet 下拉
        header = QFrame()
        header.setFrameShape(QFrame.StyledPanel)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(6, 4, 6, 4)
        header_layout.setSpacing(2)

        file_label = QLabel("未加载")
        file_label.setStyleSheet("font-weight: 600;")
        sheet_combo = QComboBox()
        sheet_combo.setToolTip(f"{title} Sheet 选择")
        header_layout.addWidget(file_label)
        header_layout.addWidget(sheet_combo)
        layout.addWidget(header)

        # 表格
        table = QTableWidget()
        layout.addWidget(table, 1)

        parent_layout.addWidget(panel, 1)
        return file_label, sheet_combo, table

    def _configure_table(self, table: QTableWidget) -> None:
        """配置 QTableWidget 的通用属性。"""
        # ExtendedSelection + SelectItems：支持 Ctrl/Shift 多选单元格，便于批量复制
        table.setSelectionMode(QTableWidget.ExtendedSelection)
        table.setSelectionBehavior(QTableWidget.SelectItems)
        table.setAlternatingRowColors(False)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        # 行号：使用 verticalHeader，可由 show_row_numbers 切换显隐
        table.verticalHeader().setVisible(self.show_row_numbers)
        table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        hheader = table.horizontalHeader()
        hheader.setStretchLastSection(True)
        table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        # 拖拽接收：实际 drop 处理由 Task 7 完成，此处先开启标志
        table.setAcceptDrops(True)
        table.setDragEnabled(False)

    # ------------------------------------------------------------------ #
    # 菜单栏
    # ------------------------------------------------------------------ #
    def _init_menubar(self) -> None:
        """构建 5 个菜单：会话/编辑/视图/工具/帮助。"""
        menubar = self.menuBar()

        # ---- 会话(S) ----
        m_session = menubar.addMenu("会话(&S)")

        self.action_open_left = QAction("打开左", self)
        self.action_open_left.setShortcut(QKeySequence("Ctrl+L"))
        self.action_open_left.triggered.connect(self._on_open_left)
        m_session.addAction(self.action_open_left)

        self.action_open_right = QAction("打开右", self)
        self.action_open_right.setShortcut(QKeySequence("Ctrl+R"))
        self.action_open_right.triggered.connect(self._on_open_right)
        m_session.addAction(self.action_open_right)

        m_session.addSeparator()

        self.action_save = QAction("保存", self)
        self.action_save.setShortcut(QKeySequence.Save)
        self.action_save.triggered.connect(self.save)
        m_session.addAction(self.action_save)

        self.action_save_as = QAction("另存为", self)
        self.action_save_as.setShortcut(QKeySequence("Ctrl+Shift+S"))
        self.action_save_as.triggered.connect(self.save_as)
        m_session.addAction(self.action_save_as)

        m_session.addSeparator()

        self.action_swap = QAction("交换左右", self)
        self.action_swap.setShortcut(QKeySequence("Ctrl+Shift+X"))
        self.action_swap.triggered.connect(self.swap_sides)
        m_session.addAction(self.action_swap)

        self.action_reload = QAction("重载文件", self)
        self.action_reload.triggered.connect(self.reload_files)
        m_session.addAction(self.action_reload)

        self.action_recompare = QAction("重新比较", self)
        self.action_recompare.setShortcut(QKeySequence("F5"))
        self.action_recompare.triggered.connect(self.recompare)
        m_session.addAction(self.action_recompare)

        # ---- 编辑(E) ----
        m_edit = menubar.addMenu("编辑(&E)")

        self.action_undo = QAction("撤销", self)
        self.action_undo.setShortcut(QKeySequence.Undo)
        self.action_undo.triggered.connect(self.undo)
        m_edit.addAction(self.action_undo)

        self.action_redo = QAction("重做", self)
        self.action_redo.setShortcut(QKeySequence.Redo)
        self.action_redo.triggered.connect(self.redo)
        m_edit.addAction(self.action_redo)

        m_edit.addSeparator()

        self.action_align_rows = QAction("对齐行", self)
        self.action_align_rows.triggered.connect(self.align_rows)
        m_edit.addAction(self.action_align_rows)

        self.action_copy_to_left = QAction("复制到左侧", self)
        self.action_copy_to_left.triggered.connect(self.copy_to_left)
        m_edit.addAction(self.action_copy_to_left)

        self.action_copy_to_right = QAction("复制到右侧", self)
        self.action_copy_to_right.triggered.connect(self.copy_to_right)
        m_edit.addAction(self.action_copy_to_right)

        m_edit.addSeparator()

        self.action_copy_cell = QAction("复制单元格", self)
        self.action_copy_cell.setShortcut(QKeySequence.Copy)
        self.action_copy_cell.triggered.connect(self.copy_cell)
        m_edit.addAction(self.action_copy_cell)

        self.action_paste = QAction("粘贴", self)
        self.action_paste.setShortcut(QKeySequence.Paste)
        self.action_paste.triggered.connect(self.paste)
        m_edit.addAction(self.action_paste)

        # ---- 视图(V) ----
        m_view = menubar.addMenu("视图(&V)")

        self.action_show_row_numbers = QAction("行号", self)
        self.action_show_row_numbers.setCheckable(True)
        self.action_show_row_numbers.setChecked(self.show_row_numbers)
        self.action_show_row_numbers.toggled.connect(self._on_toggle_row_numbers)
        m_view.addAction(self.action_show_row_numbers)

        self.action_column_settings = QAction("列设置", self)
        self.action_column_settings.triggered.connect(self.open_column_settings)
        m_view.addAction(self.action_column_settings)

        self.action_resize_columns = QAction("调整列宽", self)
        self.action_resize_columns.triggered.connect(self.resize_columns)
        m_view.addAction(self.action_resize_columns)

        self.action_show_birds_eye = QAction("差异缩略图", self)
        self.action_show_birds_eye.setCheckable(True)
        self.action_show_birds_eye.setChecked(True)
        self.action_show_birds_eye.toggled.connect(self._on_toggle_birds_eye)
        m_view.addAction(self.action_show_birds_eye)

        self.action_show_col_birds_eye = QAction("列差异缩略图", self)
        self.action_show_col_birds_eye.setCheckable(True)
        self.action_show_col_birds_eye.setChecked(True)
        self.action_show_col_birds_eye.toggled.connect(self._on_toggle_col_birds_eye)
        m_view.addAction(self.action_show_col_birds_eye)

        m_view.addSeparator()

        self.action_show_diff = QAction("显示差异", self)
        self.action_show_diff.setCheckable(True)
        self.action_show_diff.setChecked(True)
        self.action_show_diff.toggled.connect(self._on_view_filter_changed)
        m_view.addAction(self.action_show_diff)

        self.action_show_same = QAction("显示相同", self)
        self.action_show_same.setCheckable(True)
        self.action_show_same.setChecked(True)
        self.action_show_same.toggled.connect(self._on_view_filter_changed)
        m_view.addAction(self.action_show_same)

        self.action_show_left_only = QAction("显示仅左侧", self)
        self.action_show_left_only.setCheckable(True)
        self.action_show_left_only.setChecked(True)
        self.action_show_left_only.toggled.connect(self._on_view_filter_changed)
        m_view.addAction(self.action_show_left_only)

        self.action_show_right_only = QAction("显示仅右侧", self)
        self.action_show_right_only.setCheckable(True)
        self.action_show_right_only.setChecked(True)
        self.action_show_right_only.toggled.connect(self._on_view_filter_changed)
        m_view.addAction(self.action_show_right_only)

        # ---- 工具(T) ----
        m_tools = menubar.addMenu("工具(&T)")

        self.action_file_format = QAction("文件格式", self)
        self.action_file_format.triggered.connect(self.show_file_format)
        m_tools.addAction(self.action_file_format)

        self.action_options = QAction("选项", self)
        self.action_options.triggered.connect(self.show_options)
        m_tools.addAction(self.action_options)

        # ---- 帮助(H) ----
        m_help = menubar.addMenu("帮助(&H)")

        self.action_about = QAction("关于", self)
        self.action_about.triggered.connect(self.about)
        m_help.addAction(self.action_about)

        self.action_docs = QAction("文档", self)
        self.action_docs.triggered.connect(self.show_docs)
        m_help.addAction(self.action_docs)

        self.action_shortcuts = QAction("快捷键列表", self)
        self.action_shortcuts.triggered.connect(self.show_shortcuts)
        m_help.addAction(self.action_shortcuts)

    # ------------------------------------------------------------------ #
    # 工具栏
    # ------------------------------------------------------------------ #
    def _init_toolbar(self) -> None:
        """构建水平工具栏：打开左/打开右/保存/交换/上一个差异/下一个差异/统计/列设置/合并。"""
        self.toolbar = QToolBar("主工具栏", self)
        self.toolbar.setMovable(False)
        self.toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.addToolBar(self.toolbar)

        style = self.style()
        icon = lambda sp: style.standardIcon(sp)  # noqa: E731

        self.tb_open_left = self.toolbar.addAction(
            icon(QStyle.SP_DirOpenIcon), "打开左", self._on_open_left
        )
        self.tb_open_right = self.toolbar.addAction(
            icon(QStyle.SP_DirOpenIcon), "打开右", self._on_open_right
        )
        self.tb_save = self.toolbar.addAction(
            icon(QStyle.SP_DialogSaveButton), "保存", self.save
        )
        self.toolbar.addSeparator()
        self.tb_swap = self.toolbar.addAction(
            icon(QStyle.SP_BrowserReload), "交换", self.swap_sides
        )
        self.tb_prev_diff = self.toolbar.addAction(
            icon(QStyle.SP_ArrowUp), "上一个差异", self.prev_diff
        )
        self.tb_next_diff = self.toolbar.addAction(
            icon(QStyle.SP_ArrowDown), "下一个差异", self.next_diff
        )
        self.toolbar.addSeparator()
        self.tb_statistics = self.toolbar.addAction(
            icon(QStyle.SP_FileDialogListView), "统计", self.show_statistics
        )
        self.tb_column_settings = self.toolbar.addAction(
            icon(QStyle.SP_FileDialogDetailedView), "列设置", self.open_column_settings
        )
        self.tb_merge = self.toolbar.addAction(
            icon(QStyle.SP_DialogApplyButton), "合并", self.merge
        )

        # ---- 过滤器：全部 / 差异 / 相同（核心切换） ----
        self.toolbar.addSeparator()
        filter_label = QLabel("过滤器:")
        self.toolbar.addWidget(filter_label)
        self.filter_combo = QComboBox(self)
        self.filter_combo.addItem("全部")
        self.filter_combo.addItem("差异")
        self.filter_combo.addItem("相同")
        self.filter_combo.setToolTip(
            "全部：显示所有行\n"
            "差异：只显示有差异的行（隐藏完全相同的行）\n"
            "相同：只显示完全相同的行"
        )
        # 显式设置最小宽度，避免 QToolBar 压缩导致文字截断
        # 中文字符约 13-14px/字，"全部"2字 + 下拉箭头 + 边距，给足 90px
        self.filter_combo.setMinimumWidth(90)
        self.filter_combo.setSizePolicy(
            QSizePolicy.Minimum, QSizePolicy.Fixed
        )
        # 弹出下拉列表宽度与组合框一致
        self.filter_combo.view().setMinimumWidth(90)
        self.filter_combo.setCurrentIndex(0)
        self.filter_combo.currentIndexChanged.connect(
            self._on_filter_preset_changed
        )
        self.toolbar.addWidget(self.filter_combo)

    # ------------------------------------------------------------------ #
    # 状态栏
    # ------------------------------------------------------------------ #
    def _init_statusbar(self) -> None:
        """构建状态栏：就绪标签 + 行信息 + 差异数 + 备份状态 + 进度条。"""
        bar = self.statusBar()

        self.status_label = QLabel("就绪")
        bar.addWidget(self.status_label, 1)

        self.row_label = QLabel("左: 0 行 | 右: 0 行")
        bar.addPermanentWidget(self.row_label)

        self.diff_label = QLabel("差异: 0")
        bar.addPermanentWidget(self.diff_label)

        self.backup_label = QLabel("备份: 无")
        bar.addPermanentWidget(self.backup_label)

        # 比较进度条（默认隐藏，比较期间显示）
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(200)
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.hide()
        bar.addPermanentWidget(self.progress_bar)

    # ------------------------------------------------------------------ #
    # 信号连接
    # ------------------------------------------------------------------ #
    def _init_connections(self) -> None:
        """连接 Sheet 下拉、同步滚动、右键菜单、底部导航栏与单元格追踪。"""
        self.left_sheet_combo.currentIndexChanged.connect(
            lambda _: self._on_sheet_changed("left")
        )
        self.right_sheet_combo.currentIndexChanged.connect(
            lambda _: self._on_sheet_changed("right")
        )

        # 同步滚动（双向，带防递归守卫）
        self.left_table.verticalScrollBar().valueChanged.connect(
            lambda v: self._sync_scroll(self.left_table, self.right_table, "v", v)
        )
        self.right_table.verticalScrollBar().valueChanged.connect(
            lambda v: self._sync_scroll(self.right_table, self.left_table, "v", v)
        )
        self.left_table.horizontalScrollBar().valueChanged.connect(
            lambda v: self._sync_scroll(self.left_table, self.right_table, "h", v)
        )
        self.right_table.horizontalScrollBar().valueChanged.connect(
            lambda v: self._sync_scroll(self.right_table, self.left_table, "h", v)
        )

        # 右键上下文菜单
        for table in (self.left_table, self.right_table):
            table.setContextMenuPolicy(Qt.CustomContextMenu)
            table.customContextMenuRequested.connect(self.on_context_menu)

        # 单元格坐标追踪
        self.left_table.currentItemChanged.connect(
            lambda cur, prev: self._on_current_cell_changed(self.left_table, cur)
        )
        self.right_table.currentItemChanged.connect(
            lambda cur, prev: self._on_current_cell_changed(self.right_table, cur)
        )

        # 底部导航栏信号
        self.bottom_bar.sheet_activated.connect(self._on_tab_sheet_activated)
        self.bottom_bar.sheet_renamed.connect(self._on_tab_sheet_renamed)
        self.bottom_bar.sheet_close_requested.connect(self._on_tab_sheet_closed)

    def _init_shortcuts(self) -> None:
        """注册无菜单项的额外快捷键（Task 11）。"""
        # Cmd/Ctrl+O 智能打开
        QShortcut(QKeySequence("Ctrl+O"), self, activated=self._smart_open)
        # Cmd/Ctrl+Down 下一个差异 / Cmd/Ctrl+Up 上一个差异
        QShortcut(QKeySequence("Ctrl+Down"), self, activated=self.next_diff)
        QShortcut(QKeySequence("Ctrl+Up"), self, activated=self.prev_diff)

    def _init_theme_listener(self) -> None:
        """监听系统外观变化（PySide6 6.5+），自动切换主题。"""
        sh = QGuiApplication.styleHints()
        if sh is not None and hasattr(sh, "colorSchemeChanged"):
            try:
                sh.colorSchemeChanged.connect(self._on_color_scheme_changed)
            except Exception:  # noqa: BLE001
                pass

    def _on_color_scheme_changed(self) -> None:
        """系统外观变化时重新应用主题并重渲染差异。"""
        app = QApplication.instance()
        if app is not None:
            from src.gui.themes import apply_theme

            apply_theme(app)
        self._render_diffs()

    # ------------------------------------------------------------------ #
    # 文件加载（完整实现）
    # ------------------------------------------------------------------ #
    def open_left(self) -> None:
        """打开左侧文件：弹出原生文件选择对话框，加载后若两侧齐全则比较。"""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择左侧 Excel 文件", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        self.load_file("left", path)
        if self.left_path and self.right_path:
            self._run_compare()

    def open_right(self) -> None:
        """打开右侧文件：弹出原生文件选择对话框，加载后若两侧齐全则比较。"""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择右侧 Excel 文件", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        self.load_file("right", path)
        if self.left_path and self.right_path:
            self._run_compare()

    def _smart_open(self) -> None:
        """智能打开：优先打开空侧（左优先），两侧均已加载时打开左侧。"""
        if self.left_path is None:
            self.open_left()
        elif self.right_path is None:
            self.open_right()
        else:
            self.open_left()

    def _on_open_left(self) -> None:
        self.open_left()

    def _on_open_right(self) -> None:
        self.open_right()

    def load_file(self, side: str, path: str) -> None:
        """加载指定侧文件：校验格式 -> 加载工作簿 -> 填充 Sheet 下拉 -> 切换 Sheet。

        side: "left" 或 "right"。
        """
        if not str(path).lower().endswith(".xlsx"):
            QMessageBox.warning(
                self, "文件格式不支持",
                f"仅支持 .xlsx 格式文件:\n{path}\n如需支持旧版 .xls，请先另存为 .xlsx。",
            )
            return

        try:
            wb = ExcelLoader.load_workbook(path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "加载失败", f"无法加载文件:\n{path}\n\n{exc}")
            return

        sheet_names = ExcelLoader.get_sheet_names(wb)
        if not sheet_names:
            QMessageBox.warning(self, "空工作簿", f"文件中没有工作表:\n{path}")
            return

        if side == "left":
            self.left_path = path
            self.left_wb = wb
            combo = self.left_sheet_combo
        else:
            self.right_path = path
            self.right_wb = wb
            combo = self.right_sheet_combo

        # 填充下拉（屏蔽信号避免重复触发）
        combo.blockSignals(True)
        combo.clear()
        for name in sheet_names:
            combo.addItem(name)
        if sheet_names:
            combo.setCurrentIndex(0)
        combo.blockSignals(False)

        self._update_file_label(side)
        self._on_sheet_changed(side)

        # 两侧均加载后，更新底部导航栏标签与红点
        if self.left_wb and self.right_wb:
            self._refresh_bottom_bar_sheets()

        self.update_status(f"已加载 {side} 侧: {os.path.basename(path)}")

    def _update_file_label(self, side: str) -> None:
        """根据当前路径刷新标题区的文件名/大小/修改时间。"""
        path = self.left_path if side == "left" else self.right_path
        label = self.left_file_label if side == "left" else self.right_file_label
        if not path:
            label.setText("未加载")
            return
        try:
            size = os.path.getsize(path)
            mtime = os.path.getmtime(path)
            mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            label.setText(
                f"{os.path.basename(path)}  |  {self._format_size(size)}  |  {mtime_str}"
            )
        except OSError:
            label.setText(os.path.basename(path))

    @staticmethod
    def _format_size(n: float) -> str:
        """将字节数格式化为易读字符串。"""
        for unit in ("B", "KB", "MB", "GB"):
            if n < 1024:
                return f"{n:.1f} {unit}"
            n /= 1024
        return f"{n:.1f} TB"

    def reload_files(self) -> None:
        """重载左右两侧已打开的文件。"""
        if self.left_path:
            self.load_file("left", self.left_path)
        if self.right_path:
            self.load_file("right", self.right_path)
        self.update_status("已重载文件")

    def recompare(self) -> None:
        """使用当前 key_cols/ignore_cols 重新比较。"""
        self._run_compare()
        self.update_status("已重新比较")

    # ------------------------------------------------------------------ #
    # 拖拽支持（Task 7）
    # ------------------------------------------------------------------ #
    def dragEnterEvent(self, event) -> None:  # noqa: N802
        """拖入事件：仅接受含 .xlsx 文件的拖拽。"""
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(".xlsx"):
                    event.acceptProposedAction()
                    return
        event.ignore()

    def dropEvent(self, event) -> None:  # noqa: N802
        """拖放事件：根据落点位置决定加载到左侧或右侧。

        - 拖到左表上方 -> 左侧；拖到右表上方 -> 右侧；
        - 单文件且无法判定落点时：左空加载左，否则右空加载右；
        - 双文件：第一个进左，第二个进右。
        """
        paths = [
            url.toLocalFile()
            for url in event.mimeData().urls()
            if url.toLocalFile().lower().endswith(".xlsx")
        ]
        if not paths:
            event.ignore()
            return
        event.acceptProposedAction()

        if len(paths) >= 2:
            self.load_file("left", paths[0])
            self.load_file("right", paths[1])
        else:
            path = paths[0]
            # underMouse() 在拖放过程中不会更新，改用全局光标位置判断落点
            cursor_pos = QCursor.pos()
            if self.left_table.rect().contains(
                self.left_table.mapFromGlobal(cursor_pos)
            ):
                side = "left"
            elif self.right_table.rect().contains(
                self.right_table.mapFromGlobal(cursor_pos)
            ):
                side = "right"
            elif self.left_path is None:
                side = "left"
            elif self.right_path is None:
                side = "right"
            else:
                side = "left"
            self.load_file(side, path)

        if self.left_path and self.right_path:
            self._run_compare()

    # ------------------------------------------------------------------ #
    # Sheet 切换 / 表格渲染 / 比较
    # ------------------------------------------------------------------ #
    def _on_sheet_changed(self, side: str) -> None:
        """下拉切换 Sheet：提取 SheetData -> 刷新表格 -> 触发比较。"""
        combo = self.left_sheet_combo if side == "left" else self.right_sheet_combo
        if combo.count() == 0:
            return
        name = combo.currentText()
        if not name:
            return
        if not self._load_sheet_data(side, name):
            return

        # 切换 Sheet 时清除旧的差异结果，避免显示错位的对齐行
        self.diff_result = None
        self._update_birds_eye()
        self._refresh_tables()
        self._run_compare()

        # 同步底部导航栏激活标签
        self.bottom_bar.set_active(name)

    def _load_sheet_data(self, side: str, name: str) -> bool:
        """加载指定侧指定 Sheet 的数据到内存。

        返回 True 表示成功；失败时弹警告并返回 False。
        """
        wb = self.left_wb if side == "left" else self.right_wb
        if wb is None:
            return False
        try:
            ws = ExcelLoader.get_worksheet(wb, name)
            sheet_data = ExcelLoader.extract_sheet_data(ws)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(self, "Sheet 读取失败", f"无法读取工作表 {name}:\n{exc}")
            return False

        if side == "left":
            self.left_sheet_data = sheet_data
            self.left_sheet_name = name
        else:
            self.right_sheet_data = sheet_data
            self.right_sheet_name = name
        return True

    def _switch_to_sheet(self, name: str) -> None:
        """从底部标签栏切换两侧到同一 Sheet（屏蔽下拉信号避免重复比较）。"""
        self.left_sheet_combo.blockSignals(True)
        self.right_sheet_combo.blockSignals(True)
        try:
            li = self.left_sheet_combo.findText(name)
            if li >= 0 and self.left_wb is not None:
                self.left_sheet_combo.setCurrentIndex(li)
                self._load_sheet_data("left", name)
            ri = self.right_sheet_combo.findText(name)
            if ri >= 0 and self.right_wb is not None:
                self.right_sheet_combo.setCurrentIndex(ri)
                self._load_sheet_data("right", name)
        finally:
            self.left_sheet_combo.blockSignals(False)
            self.right_sheet_combo.blockSignals(False)

        self.diff_result = None
        self._update_birds_eye()
        self._refresh_tables()
        self._run_compare()
        self.bottom_bar.set_active(name)

    def _refresh_tables(self) -> None:
        """刷新左右表格内容。

        - 若 diff_result 可用且两侧数据齐全：按对齐行+对齐列渲染（两侧行数/列数一致）
        - 否则：按原始行渲染，行数取 max(左, 右) 保证视觉对齐
        """
        left = self.left_sheet_data
        right = self.right_sheet_data

        if self.diff_result and left and right:
            aligned_cols = self.diff_result.aligned_cols
            left_indices = [p.left_row for p in self.diff_result.aligned_rows]
            right_indices = [p.right_row for p in self.diff_result.aligned_rows]
        else:
            n_left = len(left.values) if left else 0
            n_right = len(right.values) if right else 0
            n = max(n_left, n_right)
            max_col = max(left.max_col if left else 0, right.max_col if right else 0)
            # 无比较结果时，构建简单的 1:1 列对齐
            aligned_cols = [
                ColPair(left_col=c, right_col=c, status="same", label="")
                for c in range(max_col)
            ]
            left_indices = [i if i < n_left else None for i in range(n)]
            right_indices = [i if i < n_right else None for i in range(n)]

        self._populate_table(self.left_table, left, left_indices, aligned_cols, "left")
        self._populate_table(self.right_table, right, right_indices, aligned_cols, "right")

    def _populate_table(
        self,
        table: QTableWidget,
        sheet_data: Optional[SheetData],
        row_indices: List[Optional[int]],
        aligned_cols: List[ColPair],
        side: str,
    ) -> None:
        """按给定的源行索引列表和对齐列填充表格。

        - row_indices 中 None 表示该显示行为虚拟空行（对应侧无此行）。
        - aligned_cols 描述列对齐：same 列两侧都有值；left_only 列右侧为虚拟空列；
          right_only 列左侧为虚拟空列。
        - 虚拟空行/空列创建占位 QTableWidgetItem（文本为空，标记 __virtual__），
          以便后续 _render_diffs 可对其着色。
        """
        table.blockSignals(True)
        try:
            table.setRowCount(0)
            n_cols = len(aligned_cols)
            table.setColumnCount(n_cols)

            # 列头：使用对齐后的列名
            labels = [cp.label if cp.label else "" for cp in aligned_cols]
            table.setHorizontalHeaderLabels(labels)

            table.setRowCount(len(row_indices))
            table.setVerticalHeaderLabels(
                [str(i + 1) for i in range(len(row_indices))]
            )

            for r, src_idx in enumerate(row_indices):
                for c, cp in enumerate(aligned_cols):
                    # 判断该单元格是否为虚拟空列
                    if side == "left":
                        col_idx = cp.left_col
                    else:
                        col_idx = cp.right_col

                    if col_idx is None:
                        # 虚拟空列：该侧无此列
                        item = QTableWidgetItem("")
                        item.setData(Qt.ItemDataRole.UserRole, "__virtual__")
                        table.setItem(r, c, item)
                        continue

                    # 虚拟空行 或 有数据
                    if src_idx is not None and sheet_data is not None:
                        values = (
                            sheet_data.values[src_idx]
                            if src_idx < len(sheet_data.values)
                            else []
                        )
                        text = values[col_idx] if col_idx < len(values) else ""
                        table.setItem(r, c, QTableWidgetItem(text))
                    else:
                        # 虚拟空行
                        item = QTableWidgetItem("")
                        item.setData(Qt.ItemDataRole.UserRole, "__virtual__")
                        table.setItem(r, c, item)

            table.verticalHeader().setVisible(self.show_row_numbers)
        finally:
            table.blockSignals(False)

    def _run_compare(self) -> None:
        """两侧数据齐全时启动后台比较线程，避免阻塞 UI。

        比较完成后由 worker 信号驱动 _on_diff_ready / _on_compare_finished
        更新表格与状态。若已有比较在跑则跳过本次请求。
        """
        if not (self.left_sheet_data and self.right_sheet_data):
            self.diff_result = None
            self._current_diff_pos = -1
            self._update_stat_labels()
            self._render_diffs()
            return

        # 防止重叠：已有比较线程在跑则跳过
        if self._compare_worker is not None and self._compare_worker.isRunning():
            return

        self._set_compare_actions_enabled(False)
        self.update_status("比较中...")
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.show()

        self._compare_worker = CompareWorker(
            self.left_sheet_data,
            self.right_sheet_data,
            self.key_cols,
            self.ignore_cols,
        )
        self._compare_worker.diff_ready.connect(self._on_diff_ready)
        self._compare_worker.progress.connect(self._on_compare_progress)
        self._compare_worker.finished_ok.connect(self._on_compare_finished)
        # 线程结束后清理对象
        self._compare_worker.finished.connect(self._compare_worker.deleteLater)
        self._compare_worker.start()

    def _set_compare_actions_enabled(self, enabled: bool) -> None:
        """比较期间禁用/恢复会触发比较的相关动作，避免重入。"""
        for act in (
            self.action_recompare,
            self.action_swap,
            self.action_reload,
            self.action_column_settings,
        ):
            if act is not None:
                act.setEnabled(enabled)

    def _on_diff_ready(self, result: DiffResult) -> None:
        """比较结果就绪：保存结果、重置游标、刷新表格与差异渲染。"""
        self.diff_result = result
        self._current_diff_pos = -1
        # 重新渲染为对齐行布局，再叠加差异标识
        self._refresh_tables()
        self._update_stat_labels()
        self._render_diffs()
        # 更新当前 Sheet 的差异红点状态
        self._update_current_sheet_diff_status(result)

    def _on_compare_progress(self, cur: int, total: int) -> None:
        """更新状态栏进度条。"""
        if total <= 0:
            self.progress_bar.setValue(100)
            return
        self.progress_bar.setValue(int(cur / total * 100))

    def _on_compare_finished(self) -> None:
        """比较结束：恢复动作可用、隐藏进度条、更新状态。"""
        self._set_compare_actions_enabled(True)
        self.progress_bar.hide()
        self._compare_worker = None
        self.update_status("就绪")

    def _diff_colors(self) -> dict:
        """根据当前主题返回差异配色字典。

        配色方案（参考 Beyond Compare）：
        - diff_row: 不同行的浅红底色（行级标识）
        - diff_cell: 差异单元格的深红底色（单元格级标识）
        - left_only_row: 仅左侧有的行的蓝色底色（左侧独占）
        - left_only_gap: 仅左侧有的行对应的右侧虚拟空行底色（浅蓝灰）
        - right_only_row: 仅右侧有的行的绿色底色（右侧独占）
        - right_only_gap: 仅右侧有的行对应的左侧虚拟空行底色（浅绿灰）
        - left_only_col: 仅左侧有的列的蓝色底色（列级，用于独占列本身）
        - right_only_col: 仅右侧有的列的绿色底色（列级，用于独占列本身）
        - col_gap: 虚拟空列底色（独占列的对侧），中性灰
        """
        dark = is_dark_mode()
        if dark:
            return {
                "diff_row": QColor("#4A2A2A"),
                "diff_cell": QColor("#B22222"),
                "left_only_row": QColor("#1A2A4A"),
                "left_only_gap": QColor("#1E2A38"),
                "right_only_row": QColor("#1A3A2A"),
                "right_only_gap": QColor("#1E2E28"),
                "left_only_col": QColor("#0C447C"),
                "right_only_col": QColor("#085041"),
                "col_gap": QColor("#252525"),
            }
        return {
            "diff_row": QColor("#FFE0E0"),
            "diff_cell": QColor("#FF6B6B"),
            "left_only_row": QColor("#D6E4FF"),
            "left_only_gap": QColor("#EEF3FB"),
            "right_only_row": QColor("#D4F0E0"),
            "right_only_gap": QColor("#EDF7F1"),
            "left_only_col": QColor("#85B7EB"),
            "right_only_col": QColor("#5DCAA5"),
            "col_gap": QColor("#F0F0F0"),
        }

    def _render_diffs(self) -> None:
        """差异可视化 - 根据 aligned_rows 和 aligned_cols 为两侧表格着色。

        配色策略（颜色鲜明，便于区分）：
        - different 行：浅红底 + 差异单元格深红
        - left_only 行：左侧蓝色（独占），右侧浅蓝灰（虚拟空行）
        - right_only 行：右侧绿色（独占），左侧浅绿灰（虚拟空行）
        - left_only 列：左侧蓝色（独占），右侧浅灰（虚拟空列）
        - right_only 列：右侧绿色（独占），左侧浅灰（虚拟空列）
        - 行级着色覆盖列级着色（行优先），但独占列的虚拟空侧保持 col_gap
        """
        colors = self._diff_colors()
        transparent = QBrush(Qt.transparent)
        n_rows = self.left_table.rowCount()

        # 1. 复位背景与行隐藏
        for table in (self.left_table, self.right_table):
            for r in range(table.rowCount()):
                table.setRowHidden(r, False)
                for c in range(table.columnCount()):
                    item = table.item(r, c)
                    if item is not None:
                        item.setBackground(transparent)

        # 2. 无比较结果时直接返回
        if not self.diff_result:
            return

        aligned = self.diff_result.aligned_rows
        aligned_cols = self.diff_result.aligned_cols
        # 视图过滤开关
        show_diff = self.action_show_diff.isChecked()
        show_same = self.action_show_same.isChecked()
        show_left_only = self.action_show_left_only.isChecked()
        show_right_only = self.action_show_right_only.isChecked()

        # 3. 列级着色：独占列本身 + 虚拟空列
        for c, cp in enumerate(aligned_cols):
            if cp.status == "left_only":
                # 左侧独占列：左侧蓝色，右侧虚拟空列灰色
                self._paint_col(self.left_table, c, colors["left_only_col"])
                self._paint_col(self.right_table, c, colors["col_gap"])
            elif cp.status == "right_only":
                # 右侧独占列：右侧绿色，左侧虚拟空列灰色
                self._paint_col(self.right_table, c, colors["right_only_col"])
                self._paint_col(self.left_table, c, colors["col_gap"])

        # 4. 逐行着色 / 过滤（行级着色覆盖列级，但保留虚拟空侧）
        for i in range(min(n_rows, len(aligned))):
            pair = aligned[i]
            status = pair.status

            # 行可见性过滤
            if status == "different" and not show_diff:
                self._hide_row_pair(i)
                continue
            if status == "same" and not show_same:
                self._hide_row_pair(i)
                continue
            if status == "left_only" and not show_left_only:
                self._hide_row_pair(i)
                continue
            if status == "right_only" and not show_right_only:
                self._hide_row_pair(i)
                continue

            # 行级着色
            if status == "different":
                # 浅红底 + 差异单元格深红
                self._paint_row(self.left_table, i, colors["diff_row"])
                self._paint_row(self.right_table, i, colors["diff_row"])
                # 差异单元格深红（覆盖行底色）
                for c in pair.diff_cells:
                    self._paint_cell(self.left_table, i, c, colors["diff_cell"])
                    self._paint_cell(self.right_table, i, c, colors["diff_cell"])
                # 独占列的虚拟空侧恢复 col_gap（覆盖行底色）
                for c, cp in enumerate(aligned_cols):
                    if cp.status == "left_only":
                        self._paint_cell(self.right_table, i, c, colors["col_gap"])
                    elif cp.status == "right_only":
                        self._paint_cell(self.left_table, i, c, colors["col_gap"])
            elif status == "left_only":
                # 左侧蓝色（独占行），右侧浅蓝灰（虚拟空行）
                self._paint_row(self.left_table, i, colors["left_only_row"])
                self._paint_row(self.right_table, i, colors["left_only_gap"])
            elif status == "right_only":
                # 右侧绿色（独占行），左侧浅绿灰（虚拟空行）
                self._paint_row(self.right_table, i, colors["right_only_row"])
                self._paint_row(self.left_table, i, colors["right_only_gap"])
            # same: 不着色（列级着色保留）

        # 同步全局差异缩略图标记
        self._update_birds_eye()

    def _update_birds_eye(self) -> None:
        """根据当前 diff_result 刷新全局差异缩略图的差异行/列标记。"""
        if self.diff_result:
            # 行方向
            row_types = [
                self.diff_result.aligned_rows[i].status
                for i in self.diff_result.diff_row_indices
            ]
            self.diff_birds_eye.set_diff_rows(
                self.diff_result.diff_row_indices, row_types
            )
            # 列方向：从 aligned_cols 收集差异列
            diff_col_indices = [
                i for i, cp in enumerate(self.diff_result.aligned_cols)
                if cp.status in ("left_only", "right_only")
            ]
            col_types = [
                self.diff_result.aligned_cols[i].status for i in diff_col_indices
            ]
            self.diff_col_birds_eye.set_diff_cols(diff_col_indices, col_types)
        else:
            self.diff_birds_eye.clear()
            self.diff_col_birds_eye.clear()

    def _hide_row_pair(self, row: int) -> None:
        """同时隐藏两侧表格的指定行。"""
        self.left_table.setRowHidden(row, True)
        self.right_table.setRowHidden(row, True)

    def _paint_row(self, table: QTableWidget, row: int, color: QColor) -> None:
        """为指定行的所有单元格设置背景色。"""
        brush = QBrush(color)
        for c in range(table.columnCount()):
            item = table.item(row, c)
            if item is not None:
                item.setBackground(brush)

    def _paint_col(self, table: QTableWidget, col: int, color: QColor) -> None:
        """为指定列的所有单元格设置背景色（用于虚拟空列）。"""
        brush = QBrush(color)
        for r in range(table.rowCount()):
            item = table.item(r, col)
            if item is not None:
                item.setBackground(brush)

    def _paint_cell(
        self, table: QTableWidget, row: int, col: int, color: QColor
    ) -> None:
        """为指定单元格设置背景色。"""
        item = table.item(row, col)
        if item is not None:
            item.setBackground(QBrush(color))

    def _update_stat_labels(self) -> None:
        """更新状态栏的行/差异数，按类型分别显示。"""
        if self.diff_result:
            s = self.diff_result.stats
            self.diff_label.setText(
                f"差异: {s['different'] + s['left_only'] + s['right_only']}"
                f" (红:{s['different']} 蓝:{s['left_only']} 绿:{s['right_only']})"
            )
        else:
            self.diff_label.setText("差异: 0")

        n_left = len(self.left_sheet_data.values) if self.left_sheet_data else 0
        n_right = len(self.right_sheet_data.values) if self.right_sheet_data else 0
        self.row_label.setText(f"左: {n_left} 行 | 右: {n_right} 行")

    # ------------------------------------------------------------------ #
    # 保存与备份（Task 9 实装）
    # ------------------------------------------------------------------ #
    def save(self) -> None:
        """保存左侧工作簿（目标文件 A）到原路径，并做完整性校验。

        流程：备份（若有未保存修改且尚未备份）-> 保存 -> 完整性校验 -> 复位脏标记。
        """
        if not (self.left_wb and self.left_path):
            QMessageBox.warning(self, "无可保存", "无可保存的文件")
            return
        if not self._dirty:
            self.update_status("无未保存更改")
        # FR-05-03：保存前自动备份原始文件（仅当有未保存修改且尚未备份时）
        self._backup_if_needed()
        try:
            ExcelLoader.save_workbook(self.left_wb, self.left_path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self, "保存失败", f"保存文件失败:\n{self.left_path}\n\n{exc}"
            )
            return

        # FR-05 可靠性：保存后完整性校验
        if not self._verify_integrity(self.left_path):
            return

        self._dirty = False
        self._backup_done = True
        if self._last_backup_path:
            self.backup_label.setText(
                f"备份: {os.path.basename(self._last_backup_path)}"
            )
        self.update_status("已保存")

    def save_as(self) -> None:
        """另存左侧工作簿到用户指定路径，不改变 left_path。

        另存前若存在未保存修改，先备份原始文件（FR-05-02：不破坏原文件）。
        """
        if not self.left_wb:
            QMessageBox.warning(self, "无可保存", "无可保存的文件")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "另存为", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        # 确保 .xlsx 后缀
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        # 有未保存修改时先备份原始文件
        self._backup_if_needed()
        try:
            ExcelLoader.save_workbook(self.left_wb, path)
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self, "另存失败", f"另存文件失败:\n{path}\n\n{exc}"
            )
            return
        if not self._verify_integrity(path):
            return
        # 注意：不修改 self.left_path，原始文件路径保持不变
        self.update_status(f"已另存为 {path}")

    def _backup_if_needed(self) -> None:
        """保存前自动备份左侧原文件（FR-05-03）。

        - left_path 为空则跳过；
        - 已备份过当前脏状态（_backup_done=True）则跳过；
        - 备份命名 ``<stem>_backup_<YYYYMMDD_HHMMSS>.xlsx``，与原文件同目录；
        - 失败仅警告，不阻断保存流程。
        """
        if not self.left_path:
            return
        if self._backup_done:
            return
        try:
            dir_name = os.path.dirname(self.left_path)
            stem = os.path.splitext(os.path.basename(self.left_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(
                dir_name, f"{stem}_backup_{timestamp}.xlsx"
            )
            shutil.copy2(self.left_path, backup_path)
            self._backup_done = True
            self._last_backup_path = backup_path
            self.backup_label.setText(f"备份: {os.path.basename(backup_path)}")
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(
                self, "备份失败",
                f"创建备份失败，保存将继续:\n{exc}",
            )

    def _verify_integrity(self, path: str) -> bool:
        """保存后完整性校验：重新加载文件确认未损坏（FR-05 可靠性）。

        成功返回 True；失败弹警告并返回 False。
        """
        try:
            ExcelLoader.load_workbook(path)
            return True
        except Exception as exc:  # noqa: BLE001
            QMessageBox.warning(
                self, "完整性校验",
                f"保存文件可能损坏: {exc}",
            )
            return False

    def _backup_before_merge(self, side: str) -> None:
        """合并前备份指定侧文件。

        side 为 "left" 或 "right"。备份文件与原文件同目录，
        命名形如 ``<stem>_backup_<YYYYMMDD_HHMMSS><ext>``。
        失败时仅警告，不中断合并流程。
        """
        path = self.left_path if side == "left" else self.right_path
        if not path:
            return
        try:
            dir_name, name = os.path.split(path)
            stem, ext = os.path.splitext(name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(
                dir_name, f"{stem}_backup_{timestamp}{ext}"
            )
            shutil.copy2(path, backup_path)
            self.backup_label.setText(f"备份: {os.path.basename(backup_path)}")
            self._backup_done = True
            self._last_backup_path = backup_path
            self.update_status(f"已创建备份 {backup_path}")
        except Exception as exc:  # noqa: BLE001
            # 备份失败不阻断合并，仅提示
            QMessageBox.warning(
                self, "备份失败",
                f"创建备份失败，合并将继续:\n{exc}",
            )

    # ------------------------------------------------------------------ #
    # 交换左右
    # ------------------------------------------------------------------ #
    def swap_sides(self) -> None:
        """交换左右两侧的路径/工作簿/数据/Sheet 名称与下拉。"""
        (
            self.left_path,
            self.right_path,
        ) = self.right_path, self.left_path
        self.left_wb, self.right_wb = self.right_wb, self.left_wb
        (
            self.left_sheet_data,
            self.right_sheet_data,
        ) = self.right_sheet_data, self.left_sheet_data
        (
            self.left_sheet_name,
            self.right_sheet_name,
        ) = self.right_sheet_name, self.left_sheet_name

        self._swap_combos()
        self._update_file_label("left")
        self._update_file_label("right")

        # 交换后差异摘要不变（只是左右对调），刷新红点
        if self.left_wb and self.right_wb:
            diff_set = {n for n, has in self._sheet_diff_summary.items() if has}
            self.bottom_bar.update_diff_status(diff_set)
            current = self.left_sheet_name or self.right_sheet_name
            if current:
                self.bottom_bar.set_active(current)

        self._refresh_tables()
        self._run_compare()
        self.update_status("已交换左右两侧")

    def _swap_combos(self) -> None:
        """交换左右 Sheet 下拉的项目与当前选择（屏蔽信号）。"""
        lc, rc = self.left_sheet_combo, self.right_sheet_combo
        lc.blockSignals(True)
        rc.blockSignals(True)
        try:
            left_items = [lc.itemText(i) for i in range(lc.count())]
            left_idx = lc.currentIndex()
            right_items = [rc.itemText(i) for i in range(rc.count())]
            right_idx = rc.currentIndex()

            lc.clear()
            lc.addItems(right_items)
            if right_items:
                lc.setCurrentIndex(min(right_idx, len(right_items) - 1))
            rc.clear()
            rc.addItems(left_items)
            if left_items:
                rc.setCurrentIndex(min(left_idx, len(left_items) - 1))
        finally:
            lc.blockSignals(False)
            rc.blockSignals(False)

    # ------------------------------------------------------------------ #
    # 差异导航（Task 6 实装）
    # ------------------------------------------------------------------ #
    def next_diff(self) -> None:
        """跳转到下一个差异行（循环），并在两侧表格同步选中与滚动。"""
        indices = self.diff_result.diff_row_indices if self.diff_result else []
        if not indices:
            self.update_status("无差异")
            return
        self._current_diff_pos = (self._current_diff_pos + 1) % len(indices)
        self._goto_diff(indices[self._current_diff_pos], len(indices))

    def prev_diff(self) -> None:
        """跳转到上一个差异行（循环），并在两侧表格同步选中与滚动。"""
        indices = self.diff_result.diff_row_indices if self.diff_result else []
        if not indices:
            self.update_status("无差异")
            return
        if self._current_diff_pos <= 0:
            self._current_diff_pos = len(indices) - 1
        else:
            self._current_diff_pos -= 1
        self._goto_diff(indices[self._current_diff_pos], len(indices))

    def _goto_diff(self, target_row: int, total: int) -> None:
        """选中并滚动两侧表格至指定对齐行，更新状态栏。"""
        for table in (self.left_table, self.right_table):
            if 0 <= target_row < table.rowCount():
                table.setCurrentCell(target_row, 0)
                item = table.item(target_row, 0)
                if item is not None:
                    table.scrollToItem(item)
        pos = self._current_diff_pos + 1
        self.update_status(f"差异 {pos}/{total}")

    def show_statistics(self) -> None:
        """弹窗展示当前比较统计。"""
        if not self.diff_result:
            QMessageBox.information(self, "统计", "尚未进行比较，无统计数据。")
            return
        s = self.diff_result.stats
        total = len(self.diff_result.aligned_rows)
        info = (
            f"对齐行总数: {total}\n"
            f"相同: {s['same']}\n"
            f"不同: {s['different']}\n"
            f"仅左: {s['left_only']}\n"
            f"仅右: {s['right_only']}\n"
            f"差异单元格: {len(self.diff_result.diff_cell_set)}"
        )
        QMessageBox.information(self, "差异统计", info)

    # ------------------------------------------------------------------ #
    # 列设置 / 合并 / 上下文菜单（Task 8 实装）
    # ------------------------------------------------------------------ #
    def open_column_settings(self) -> None:
        """打开列设置对话框，配置 Key 列与忽略列后触发重新比较。"""
        if not self.left_sheet_data and not self.right_sheet_data:
            QMessageBox.warning(self, "列设置", "请先加载文件")
            return
        # 列标签：取两侧列数较多者，保证覆盖所有列
        left_labels = (
            self.left_sheet_data.header_labels if self.left_sheet_data else []
        )
        right_labels = (
            self.right_sheet_data.header_labels if self.right_sheet_data else []
        )
        if len(left_labels) >= len(right_labels):
            labels = left_labels
        else:
            labels = right_labels

        dlg = ColumnSettingsDialog(labels, self.key_cols, self.ignore_cols, self)
        if dlg.exec() == QDialog.Accepted:
            self.key_cols = dlg.get_key_cols()
            self.ignore_cols = dlg.get_ignore_cols()
            self.update_status(
                f"Key 列: {self.key_cols} | 忽略列: {self.ignore_cols}"
            )
            self._run_compare()

    def merge(self) -> None:
        """弹出合并策略选择菜单。"""
        if not self.diff_result:
            self.update_status("无差异结果，无法合并")
            return
        menu = QMenu(self)
        menu.addAction("右覆盖（以右侧为准）", self.merge_right_wins)
        menu.addAction("左覆盖（以左侧为准）", self.merge_left_wins)
        menu.addAction("追加差异行", self.merge_append)
        menu.addAction("手动合并", self.merge_manual)

        btn = self.sender()
        if isinstance(btn, QAction):
            # 由工具栏 QAction 触发：在工具栏按钮下方弹出
            widget = self.toolbar.widgetForAction(btn)
            if widget is not None:
                pos = widget.mapToGlobal(widget.rect().bottomLeft())
                menu.exec(pos)
                return
        menu.exec(QCursor.pos())

    def merge_right_wins(self) -> None:
        """右覆盖：以右侧为准，将右侧差异覆盖到左侧。"""
        if not self.diff_result:
            self.update_status("无差异结果，无法合并")
            return
        btn = QMessageBox.question(
            self, "右覆盖", "将以右侧为准覆盖左侧差异，继续？"
        )
        if btn != QMessageBox.Yes:
            return
        try:
            ExcelMerger.merge_right_to_left(
                self.diff_result, self.left_sheet_data, self.right_sheet_data
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "合并失败", f"右覆盖合并失败:\n{exc}")
            return
        self._post_merge("left")
        self.update_status("已合并（右覆盖）— 请保存")

    def merge_left_wins(self) -> None:
        """左覆盖：以左侧为准，将左侧差异覆盖到右侧。"""
        if not self.diff_result:
            self.update_status("无差异结果，无法合并")
            return
        btn = QMessageBox.question(
            self, "左覆盖", "将以左侧为准覆盖右侧差异，继续？"
        )
        if btn != QMessageBox.Yes:
            return
        try:
            ExcelMerger.merge_left_to_right(
                self.diff_result, self.left_sheet_data, self.right_sheet_data
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "合并失败", f"左覆盖合并失败:\n{exc}")
            return
        self._post_merge("right")
        self.update_status("已合并（左覆盖）— 请保存")

    def merge_append(self) -> None:
        """追加差异行：仅把右侧独占行追加到左侧末尾。"""
        if not self.diff_result:
            self.update_status("无差异结果，无法合并")
            return
        btn = QMessageBox.question(
            self, "追加差异行", "将把右侧独占行追加到左侧末尾，继续？"
        )
        if btn != QMessageBox.Yes:
            return
        try:
            ExcelMerger.append_rows(
                self.diff_result,
                self.left_sheet_data,
                self.right_sheet_data,
                self.key_cols,
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(self, "合并失败", f"追加差异行失败:\n{exc}")
            return
        self._post_merge("left")
        self.update_status("已合并（追加差异行）— 请保存")

    def merge_manual(self) -> None:
        """手动合并：提示用户通过右键菜单逐项合并。"""
        self.update_status("请使用右键菜单逐项合并")

    def _post_merge(self, side: str) -> None:
        """合并后处理：重新提取数据 -> 刷新表格 -> 重新比较 -> 标记未保存。

        side 为合并写入的目标侧（"left" 或 "right"）。由于合并器就地修改了
        worksheet 对象，需重新提取 SheetData 以反映最新值。
        """
        if side == "left" and self.left_sheet_data is not None:
            self.left_sheet_data = ExcelLoader.extract_sheet_data(
                self.left_sheet_data.worksheet
            )
        elif side == "right" and self.right_sheet_data is not None:
            self.right_sheet_data = ExcelLoader.extract_sheet_data(
                self.right_sheet_data.worksheet
            )
        # 清除旧差异结果，避免按过期对齐渲染
        self.diff_result = None
        self._update_birds_eye()
        self._refresh_tables()
        self._run_compare()
        self._dirty = True
        self._backup_done = False
        self.backup_label.setText("备份: 待保存")
        # 合并后重新计算各 sheet 差异红点（可能当前 sheet 差异已消除）
        if self.left_wb and self.right_wb:
            self._compute_sheet_diff_summary()
            diff_set = {n for n, has in self._sheet_diff_summary.items() if has}
            self.bottom_bar.update_diff_status(diff_set)

    def on_context_menu(self, pos) -> None:
        """表格右键菜单：根据触发表格决定“复制到对侧”的文案与方向。

        记录上下文（_ctx_table/_ctx_row/_ctx_col）供 _ctx_* 回调读取。
        """
        table = self.sender()
        if table is None:
            return
        self._ctx_table = table
        self._ctx_row = table.currentRow()
        self._ctx_col = table.currentColumn()

        is_left = table is self.left_table
        other_label = "右侧" if is_left else "左侧"
        menu = QMenu(self)
        menu.addAction(f"复制行到{other_label}（含插入/删除）", self._ctx_copy_row)
        menu.addAction(f"复制列到{other_label}（含插入/删除）", self._ctx_copy_column)
        menu.addAction(f"复制单元格到{other_label}", self._ctx_copy_single_cell)
        menu.addSeparator()
        menu.addAction("复制单元格值", self._ctx_copy_cell)
        menu.addSeparator()
        menu.addAction("交换左右", self.swap_sides)
        global_pos = table.viewport().mapToGlobal(pos)
        menu.exec(global_pos)

    def _resolve_pair_for_row(self, row: int, quiet: bool = False):
        """根据 _ctx_table 与指定行解析对齐行，返回 (source, src_idx, target, tgt_idx, side, mode)。

        - mode="overwrite": target 行已存在，覆盖复制（双侧均有该行）。
        - mode="insert": target 需在该位置插入新行后复制（用于 left_only/right_only 行，
          即在多的一侧右键，对侧对应位置是虚拟空行的情况）。tgt_idx 为 target 中 0-based 插入位置。
        - mode="delete": 当前侧是虚拟空行（少的一侧），需删除对侧多出的行。
          tgt_idx 为对侧 0-based 行索引。
        - 无法操作（无差异结果、行越界）时返回 None。
        quiet=True 时不更新状态栏（用于批量场景静默跳过）。
        """
        if not self.diff_result or row < 0:
            if not quiet:
                self.update_status("无有效行可操作")
            return None
        if row >= len(self.diff_result.aligned_rows):
            if not quiet:
                self.update_status("行索引越界")
            return None
        pair = self.diff_result.aligned_rows[row]
        aligned = self.diff_result.aligned_rows
        is_left = self._ctx_table is self.left_table

        if is_left:
            # 从左侧操作
            if pair.left_row is None:
                # 左侧是虚拟空行（right_only 行），右侧多出 -> 删除右侧对应行
                return (
                    None, None,
                    self.right_sheet_data, pair.right_row,
                    "right", "delete",
                )
            if pair.right_row is not None:
                # 双侧都有该行 -> 覆盖
                return (
                    self.left_sheet_data,
                    pair.left_row,
                    self.right_sheet_data,
                    pair.right_row,
                    "right",
                    "overwrite",
                )
            # left_only 行：右侧对应位置是虚拟空行 -> 插入到右侧
            insert_idx = sum(1 for p in aligned[:row] if p.right_row is not None)
            return (
                self.left_sheet_data,
                pair.left_row,
                self.right_sheet_data,
                insert_idx,
                "right",
                "insert",
            )
        else:
            # 从右侧操作
            if pair.right_row is None:
                # 右侧是虚拟空行（left_only 行），左侧多出 -> 删除左侧对应行
                return (
                    None, None,
                    self.left_sheet_data, pair.left_row,
                    "left", "delete",
                )
            if pair.left_row is not None:
                return (
                    self.right_sheet_data,
                    pair.right_row,
                    self.left_sheet_data,
                    pair.left_row,
                    "left",
                    "overwrite",
                )
            # right_only 行：左侧对应位置是虚拟空行 -> 插入到左侧
            insert_idx = sum(1 for p in aligned[:row] if p.left_row is not None)
            return (
                self.right_sheet_data,
                pair.right_row,
                self.left_sheet_data,
                insert_idx,
                "left",
                "insert",
            )

    def _ctx_resolve_pair(self):
        """根据 _ctx_table/_ctx_row 解析当前对齐行（单行场景保留的旧接口）。"""
        return self._resolve_pair_for_row(self._ctx_row)

    def _ctx_copy_row(self) -> None:
        """右键：把选中行整行（值与样式）复制到对侧对应行（支持多选）。

        - 覆盖模式（双侧都有该行）：直接复制到对侧已存在行。
        - 插入模式（left_only/right_only 行）：在对侧对应位置插入新行后复制，
          多个插入按从后往前执行以避免位置偏移。
        - 列映射：按 aligned_cols 的 same 列映射复制，避免列结构不同时错位。
        """
        if self._ctx_table is None:
            return
        if not self.diff_result:
            self.update_status("无差异结果，无法复制")
            return

        # 收集选中行（去重 + 排序），无选中时退化为当前行
        rows = {it.row() for it in self._ctx_table.selectedItems()}
        if not rows and self._ctx_row >= 0:
            rows = {self._ctx_row}
        if not rows:
            self.update_status("无有效行可复制")
            return

        # 先解析所有可复制行（_post_merge 会清空 diff_result，必须提前收集）
        tasks = []
        skipped = 0
        for row in sorted(rows):
            resolved = self._resolve_pair_for_row(row, quiet=True)
            if resolved is None:
                skipped += 1
                continue
            tasks.append(resolved)

        if not tasks:
            self.update_status("选中行无有效数据可复制")
            return

        side = tasks[0][4]
        is_left = self._ctx_table is self.left_table
        # 构建列映射：same 列的 (src_col, tgt_col)，按对齐顺序
        aligned_cols = self.diff_result.aligned_cols
        col_mapping: List[Tuple[int, int]] = []
        for cp in aligned_cols:
            if cp.status == "same":
                if is_left:
                    col_mapping.append((cp.left_col, cp.right_col))
                else:
                    col_mapping.append((cp.right_col, cp.left_col))

        overwrite_tasks = [t for t in tasks if t[5] == "overwrite"]
        insert_tasks = [t for t in tasks if t[5] == "insert"]
        delete_tasks = [t for t in tasks if t[5] == "delete"]

        failures = 0
        last_exc: Optional[Exception] = None

        # 先执行覆盖（不改变行数，位置稳定）
        for source, src_idx, target, tgt_idx, _side, _mode in overwrite_tasks:
            try:
                ExcelMerger.copy_row_to_other(
                    source, src_idx, target, tgt_idx, 0, col_mapping
                )
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        # 再执行插入（从后往前，避免位置偏移）
        for source, src_idx, target, tgt_idx, _side, _mode in sorted(
            insert_tasks, key=lambda t: t[3], reverse=True
        ):
            try:
                ExcelMerger.insert_row_to_other(
                    source, src_idx, target, tgt_idx, 0, col_mapping
                )
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        # 最后执行删除（从后往前，避免位置偏移）
        for _source, _src_idx, target, tgt_idx, _side, _mode in sorted(
            delete_tasks, key=lambda t: t[3], reverse=True
        ):
            try:
                ExcelMerger.delete_row(target, tgt_idx)
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        if failures:
            QMessageBox.critical(
                self, "复制失败", f"复制行失败 ({failures}/{len(tasks)}):\n{last_exc}"
            )
            return

        self._post_merge(side)
        msg = f"已复制 {len(tasks)} 行到对侧"
        if insert_tasks:
            msg += f"（其中 {len(insert_tasks)} 行为新增插入）"
        if delete_tasks:
            msg += f"（其中 {len(delete_tasks)} 行为删除对侧多出行）"
        if skipped:
            msg += f"（跳过 {skipped} 行无数据）"
        self.update_status(msg)

    def _resolve_pair_for_col(self, col: int, quiet: bool = False):
        """根据 _ctx_table 与指定列(aligned_col 索引)解析对齐列。

        返回 (source, src_col, target, tgt_col, side, mode)：
        - mode="overwrite": target 列已存在，覆盖复制（same 列）。
        - mode="insert": target 需在该位置插入新列后复制（left_only/right_only 列，
          即对侧对应位置是虚拟空列）。tgt_col 为 target 中 0-based 插入位置。
        - mode="delete": 当前侧是虚拟空列（少的一侧），需删除对侧多出的列。
          tgt_col 为对侧 0-based 列索引。
        - 无法操作时返回 None。
        """
        if not self.diff_result or col < 0:
            if not quiet:
                self.update_status("无有效列可操作")
            return None
        aligned_cols = self.diff_result.aligned_cols
        if col >= len(aligned_cols):
            if not quiet:
                self.update_status("列索引越界")
            return None
        cp = aligned_cols[col]
        is_left = self._ctx_table is self.left_table

        if is_left:
            # 从左侧操作
            if cp.left_col is None:
                # 左侧是虚拟空列（right_only 列），右侧多出 -> 删除右侧对应列
                return (
                    None, None,
                    self.right_sheet_data, cp.right_col,
                    "right", "delete",
                )
            if cp.right_col is not None:
                # same 列 -> 覆盖
                return (
                    self.left_sheet_data, cp.left_col,
                    self.right_sheet_data, cp.right_col,
                    "right", "overwrite",
                )
            # left_only 列：右侧对应位置是虚拟空列 -> 插入到右侧
            insert_idx = sum(
                1 for c in aligned_cols[:col] if c.right_col is not None
            )
            return (
                self.left_sheet_data, cp.left_col,
                self.right_sheet_data, insert_idx,
                "right", "insert",
            )
        else:
            # 从右侧操作
            if cp.right_col is None:
                # 右侧是虚拟空列（left_only 列），左侧多出 -> 删除左侧对应列
                return (
                    None, None,
                    self.left_sheet_data, cp.left_col,
                    "left", "delete",
                )
            if cp.left_col is not None:
                return (
                    self.right_sheet_data, cp.right_col,
                    self.left_sheet_data, cp.left_col,
                    "left", "overwrite",
                )
            # right_only 列：左侧对应位置是虚拟空列 -> 插入到左侧
            insert_idx = sum(
                1 for c in aligned_cols[:col] if c.left_col is not None
            )
            return (
                self.right_sheet_data, cp.right_col,
                self.left_sheet_data, insert_idx,
                "left", "insert",
            )

    def _ctx_copy_column(self) -> None:
        """右键：把选中列整列（值与样式）复制到对侧对应列（支持多选）。

        - 覆盖模式（same 列）：直接复制到对侧已存在列。
        - 插入模式（left_only/right_only 列）：在对侧对应位置插入新列后复制，
          多个插入按从后往前执行以避免位置偏移。
        - 行映射：按 aligned_rows 的 same 行映射复制，避免行结构不同时错位。
        """
        if self._ctx_table is None:
            return
        if not self.diff_result:
            self.update_status("无差异结果，无法复制")
            return

        # 收集选中列（去重 + 排序），无选中时退化为当前列
        cols = {it.column() for it in self._ctx_table.selectedItems()}
        if not cols and self._ctx_col >= 0:
            cols = {self._ctx_col}
        if not cols:
            self.update_status("无有效列可复制")
            return

        # 先解析所有可复制列（_post_merge 会清空 diff_result，必须提前收集）
        tasks = []
        skipped = 0
        for col in sorted(cols):
            resolved = self._resolve_pair_for_col(col, quiet=True)
            if resolved is None:
                skipped += 1
                continue
            tasks.append(resolved)

        if not tasks:
            self.update_status("选中列无有效数据可复制")
            return

        side = tasks[0][4]
        is_left = self._ctx_table is self.left_table
        # 构建行映射：双侧都有数据的行（same + different）的 (src_row, tgt_row)，
        # 按对齐顺序。different 行两侧都有该行，列值应当被复制，否则差异行
        # 的单元格不会被更新，导致复制后差异仍在。
        aligned_rows = self.diff_result.aligned_rows
        row_mapping: List[Tuple[int, int]] = []
        for pair in aligned_rows:
            if pair.status in ("same", "different"):
                if is_left:
                    row_mapping.append((pair.left_row, pair.right_row))
                else:
                    row_mapping.append((pair.right_row, pair.left_row))

        overwrite_tasks = [t for t in tasks if t[5] == "overwrite"]
        insert_tasks = [t for t in tasks if t[5] == "insert"]
        delete_tasks = [t for t in tasks if t[5] == "delete"]

        failures = 0
        last_exc: Optional[Exception] = None

        # 先执行覆盖（不改变列数，位置稳定）
        for source, src_col, target, tgt_col, _side, _mode in overwrite_tasks:
            try:
                ExcelMerger.copy_column_to_other(
                    source, src_col, target, tgt_col, 0, row_mapping
                )
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        # 再执行插入（从后往前，避免位置偏移）
        for source, src_col, target, tgt_col, _side, _mode in sorted(
            insert_tasks, key=lambda t: t[3], reverse=True
        ):
            try:
                ExcelMerger.insert_column_to_other(
                    source, src_col, target, tgt_col, 0, row_mapping
                )
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        # 最后执行删除（从后往前，避免位置偏移）
        for _source, _src_col, target, tgt_col, _side, _mode in sorted(
            delete_tasks, key=lambda t: t[3], reverse=True
        ):
            try:
                ExcelMerger.delete_column(target, tgt_col)
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        if failures:
            QMessageBox.critical(
                self, "复制失败", f"复制列失败 ({failures}/{len(tasks)}):\n{last_exc}"
            )
            return

        self._post_merge(side)
        msg = f"已复制 {len(tasks)} 列到对侧"
        if insert_tasks:
            msg += f"（其中 {len(insert_tasks)} 列为新增插入）"
        if delete_tasks:
            msg += f"（其中 {len(delete_tasks)} 列为删除对侧多出列）"
        if skipped:
            msg += f"（跳过 {skipped} 列无数据）"
        self.update_status(msg)

    def _ctx_copy_single_cell(self) -> None:
        """右键：把选中单元格（值与样式）复制到对侧对应位置（支持多选）。

        仅处理覆盖模式（双侧都有该行）。插入模式（left_only/right_only 行）
        不支持单格复制，会跳过并提示。
        """
        if self._ctx_table is None:
            return
        if not self.diff_result:
            self.update_status("无差异结果，无法复制")
            return

        # 收集选中 (row, col)（去重），无选中时退化为当前单元格
        cells = {(it.row(), it.column()) for it in self._ctx_table.selectedItems()}
        if not cells and self._ctx_row >= 0 and self._ctx_col >= 0:
            cells = {(self._ctx_row, self._ctx_col)}
        if not cells:
            self.update_status("无有效单元格可复制")
            return

        # 按行解析一次，避免重复 resolve
        rows_needed = {r for r, _ in cells}
        row_resolved: dict = {}
        skipped_rows = 0
        for row in rows_needed:
            res = self._resolve_pair_for_row(row, quiet=True)
            if res is None:
                skipped_rows += 1
            elif res[5] in ("insert", "delete"):
                # 插入/删除模式（对侧无该行）不支持单格复制，跳过
                skipped_rows += 1
            else:
                row_resolved[row] = res

        if not row_resolved:
            self.update_status("选中单元格所在行无有效对应行，无法复制")
            return

        side = next(iter(row_resolved.values()))[4]
        aligned_cols = self.diff_result.aligned_cols
        is_left = self._ctx_table is self.left_table
        failures = 0
        success = 0
        last_exc: Optional[Exception] = None
        for (row, col) in cells:
            resolved = row_resolved.get(row)
            if resolved is None:
                continue
            source, src_idx, target, tgt_idx, _side, _mode = resolved
            # col 是 aligned_col 索引，需映射到实际列索引
            if col >= len(aligned_cols):
                continue
            cp = aligned_cols[col]
            if cp.status != "same":
                # 虚拟空列不支持单格复制
                continue
            src_col = cp.left_col if is_left else cp.right_col
            tgt_col = cp.right_col if is_left else cp.left_col
            try:
                ExcelMerger.copy_single_cell(source, src_idx, src_col, target, tgt_idx)
                success += 1
            except Exception as exc:  # noqa: BLE001
                failures += 1
                last_exc = exc

        if failures:
            QMessageBox.critical(
                self, "复制失败", f"复制单元格失败 ({failures}):\n{last_exc}"
            )
            return

        self._post_merge(side)
        msg = f"已复制 {success} 个单元格到对侧"
        if skipped_rows:
            msg += f"（跳过 {skipped_rows} 行无对应行）"
        self.update_status(msg)

    def _ctx_copy_cell(self) -> None:
        """右键：复制选中单元格文本到系统剪贴板（支持多选）。"""
        if self._ctx_table is None:
            return
        self._copy_selected_cells(self._ctx_table)

    # ------------------------------------------------------------------ #
    # 编辑类操作（多为桩）
    # ------------------------------------------------------------------ #
    def undo(self) -> None:
        self.update_status("TODO: 撤销（Task 8/9）")

    def redo(self) -> None:
        self.update_status("TODO: 重做（Task 8/9）")

    def align_rows(self) -> None:
        self.update_status("TODO: 对齐行（Task 8）")

    def copy_to_left(self) -> None:
        self.update_status("TODO: 复制到左侧（Task 8）")

    def copy_to_right(self) -> None:
        self.update_status("TODO: 复制到右侧（Task 8）")

    def copy_cell(self) -> None:
        """复制选中单元格到剪贴板（支持多选）。

        - 多选时按行列整理为 TSV（制表符分隔列，换行分隔行），
          可直接粘贴到 Excel/WPS。
        - 单选时退化为纯文本。
        - 无选中项时优先取当前单元格。
        """
        table = self._focused_table()
        if table is None:
            self.update_status("无活动表格")
            return
        self._copy_selected_cells(table)

    def _focused_table(self) -> Optional[QTableWidget]:
        """返回当前聚焦的表格；若都无焦点则取有选中项的表格。"""
        for table in (self.left_table, self.right_table):
            if table.hasFocus():
                return table
        for table in (self.left_table, self.right_table):
            if table.selectedItems():
                return table
        return None

    def _copy_selected_cells(self, table: QTableWidget) -> None:
        """把表格中选中单元格整理为 TSV 文本写入剪贴板。"""
        items = table.selectedItems()
        if not items:
            # 退化到当前单元格
            row, col = table.currentRow(), table.currentColumn()
            if row < 0 or col < 0:
                self.update_status("无单元格值可复制")
                return
            item = table.item(row, col)
            text = item.text() if item is not None else ""
            QApplication.clipboard().setText(text)
            self.update_status("已复制单元格值")
            return

        # 收集行列范围，构建稀疏字典 {row: {col: text}}
        row_cols: dict = {}
        min_row = min(it.row() for it in items)
        max_row = max(it.row() for it in items)
        min_col = min(it.column() for it in items)
        max_col = max(it.column() for it in items)

        grid: dict = {}
        for it in items:
            grid.setdefault(it.row(), {})[it.column()] = it.text()

        # 按行拼接：缺失单元格留空
        lines = []
        for r in range(min_row, max_row + 1):
            row_dict = grid.get(r, {})
            cells = [row_dict.get(c, "") for c in range(min_col, max_col + 1)]
            lines.append("\t".join(cells))
        text = "\n".join(lines)

        QApplication.clipboard().setText(text)
        count = len(items)
        if count == 1:
            self.update_status("已复制单元格值")
        else:
            self.update_status(
                f"已复制 {count} 个单元格 ({max_row - min_row + 1}行 × "
                f"{max_col - min_col + 1}列)"
            )

    def paste(self) -> None:
        self.update_status("TODO: 粘贴（Task 8）")

    # ------------------------------------------------------------------ #
    # 视图类操作
    # ------------------------------------------------------------------ #
    def _on_toggle_row_numbers(self, checked: bool) -> None:
        self.show_row_numbers = checked
        self.left_table.verticalHeader().setVisible(checked)
        self.right_table.verticalHeader().setVisible(checked)

    def _on_toggle_birds_eye(self, checked: bool) -> None:
        """显示/隐藏全局差异缩略图导航条（行方向）。"""
        self.diff_birds_eye.setVisible(checked)

    def _on_toggle_col_birds_eye(self, checked: bool) -> None:
        """显示/隐藏全局差异缩略图导航条（列方向）。"""
        self.diff_col_birds_eye.setVisible(checked)

    def _on_view_filter_changed(self) -> None:
        """显示差异/相同/仅左/仅右 过滤变化 —— 触发重渲染（Task 6 实装过滤）。"""
        self._sync_filter_combo_from_checks()
        self._render_diffs()

    def _on_filter_preset_changed(self, idx: int) -> None:
        """工具栏过滤器预设切换：根据预设同步 4 个细粒度过滤开关。

        - 全部(0)：4 个开关全开
        - 差异(1)：显示差异/仅左/仅右，隐藏相同
        - 相同(2)：仅显示相同，隐藏差异/仅左/仅右
        """
        if idx == 0:  # 全部
            show_diff, show_same, show_left, show_right = True, True, True, True
        elif idx == 1:  # 差异
            show_diff, show_same, show_left, show_right = True, False, True, True
        else:  # 相同
            show_diff, show_same, show_left, show_right = False, True, False, False

        for act, val in (
            (self.action_show_diff, show_diff),
            (self.action_show_same, show_same),
            (self.action_show_left_only, show_left),
            (self.action_show_right_only, show_right),
        ):
            act.blockSignals(True)
            act.setChecked(val)
            act.blockSignals(False)
        self._render_diffs()

    def _sync_filter_combo_from_checks(self) -> None:
        """根据 4 个细粒度过滤开关的实际状态，反向同步工具栏过滤器预设。

        状态匹配某预设时同步显示；不匹配时保留当前选择不动，避免误导。
        """
        d = self.action_show_diff.isChecked()
        s = self.action_show_same.isChecked()
        l = self.action_show_left_only.isChecked()
        r = self.action_show_right_only.isChecked()

        if d and s and l and r:
            idx = 0  # 全部
        elif d and not s and l and r:
            idx = 1  # 差异
        elif not d and s and not l and not r:
            idx = 2  # 相同
        else:
            return  # 自定义组合，不更新预设
        self.filter_combo.blockSignals(True)
        self.filter_combo.setCurrentIndex(idx)
        self.filter_combo.blockSignals(False)

    def resize_columns(self) -> None:
        """按内容自动调整列宽。"""
        for table in (self.left_table, self.right_table):
            table.resizeColumnsToContents()
            # 末列保持拉伸
            table.horizontalHeader().setStretchLastSection(True)

    # ------------------------------------------------------------------ #
    # 工具 / 帮助
    # ------------------------------------------------------------------ #
    def show_file_format(self) -> None:
        QMessageBox.information(
            self, "文件格式",
            "本工具仅支持 Microsoft Excel 2007+ 的 .xlsx 格式。\n"
            "旧版 .xls 文件请先在 Excel 中另存为 .xlsx。",
        )

    def show_options(self) -> None:
        self.update_status("TODO: 选项")

    def about(self) -> None:
        QMessageBox.about(
            self, "关于",
            "表格比较与合并工具\n\n"
            "基于 PySide6 + openpyxl 实现 Excel 工作表的智能比较与合并。\n"
            "支持差异可视化、多策略合并、合并单元格与公式保留。",
        )

    def show_docs(self) -> None:
        self.update_status("TODO: 文档")

    def show_shortcuts(self) -> None:
        QMessageBox.information(
            self, "快捷键列表",
            "打开左: Ctrl+L\n"
            "打开右: Ctrl+R\n"
            "保存: Cmd/Ctrl+S\n"
            "另存为: Ctrl+Shift+S\n"
            "交换左右: Ctrl+Shift+X\n"
            "重新比较: F5\n"
            "撤销: Cmd/Ctrl+Z\n"
            "重做: Cmd/Ctrl+Shift+Z\n"
            "复制单元格: Cmd/Ctrl+C\n"
            "粘贴: Cmd/Ctrl+V",
        )

    # ------------------------------------------------------------------ #
    # 同步滚动
    # ------------------------------------------------------------------ #
    def _sync_scroll(self, source: QTableWidget, target: QTableWidget, axis: str, value: int) -> None:
        """同步另一侧表格的滚动条位置（带防递归守卫）。"""
        if self._sync_scroll_blocked:
            return
        self._sync_scroll_blocked = True
        try:
            if axis == "v":
                target.verticalScrollBar().setValue(value)
            else:
                target.horizontalScrollBar().setValue(value)
        finally:
            self._sync_scroll_blocked = False

    # ------------------------------------------------------------------ #
    # 底部导航栏：坐标追踪 / 标签交互 / 差异红点
    # ------------------------------------------------------------------ #
    def _on_current_cell_changed(self, table: QTableWidget, item) -> None:
        """当前单元格变化时，更新底部导航栏的坐标显示。"""
        if item is None:
            self.bottom_bar.set_coord("-")
            return
        row = item.row()
        col = item.column()
        if row < 0 or col < 0:
            self.bottom_bar.set_coord("-")
            return
        # Excel 风格坐标：列字母 + 1-based 行号
        col_letter = get_column_letter(col + 1)
        self.bottom_bar.set_coord(f"{col_letter}{row + 1}")

    def _on_tab_sheet_activated(self, name: str) -> None:
        """底部标签点击 → 切换两侧到同一 Sheet。"""
        self._switch_to_sheet(name)

    def _on_tab_sheet_renamed(self, old_name: str, new_name: str) -> None:
        """底部标签双击重命名 → 同步到两侧 workbook 与下拉。"""
        renamed_left = False
        renamed_right = False

        # 左侧 workbook
        if self.left_wb is not None and old_name in self.left_wb.sheetnames:
            try:
                self.left_wb[old_name].title = new_name
                renamed_left = True
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "重命名失败", f"左侧重命名失败:\n{exc}")

        # 右侧 workbook
        if self.right_wb is not None and old_name in self.right_wb.sheetnames:
            try:
                self.right_wb[old_name].title = new_name
                renamed_right = True
            except Exception as exc:  # noqa: BLE001
                QMessageBox.warning(self, "重命名失败", f"右侧重命名失败:\n{exc}")

        # 更新下拉
        for combo in (self.left_sheet_combo, self.right_sheet_combo):
            idx = combo.findText(old_name)
            if idx >= 0:
                combo.blockSignals(True)
                combo.setItemText(idx, new_name)
                combo.blockSignals(False)

        # 更新当前 sheet 名状态
        if self.left_sheet_name == old_name and renamed_left:
            self.left_sheet_name = new_name
        if self.right_sheet_name == old_name and renamed_right:
            self.right_sheet_name = new_name

        # 更新差异摘要缓存中的键
        if old_name in self._sheet_diff_summary:
            self._sheet_diff_summary[new_name] = self._sheet_diff_summary.pop(old_name)

        # 标记脏状态（重命名未保存到文件）
        self._dirty = True
        self._backup_done = False
        self.backup_label.setText("备份: 待保存")
        self.update_status(f"已重命名: {old_name} → {new_name}（仅本地）")

    def _on_tab_sheet_closed(self, name: str, close_others: bool) -> None:
        """底部标签右键关闭：从标签栏移除（不影响实际 workbook）。"""
        if close_others:
            for n in list(self.bottom_bar.get_sheet_names()):
                if n != name:
                    self.bottom_bar.remove_sheet(n)
            self.update_status(f"已关闭其他标签（保留 {name}）")
        else:
            self.bottom_bar.remove_sheet(name)
            self.update_status(f"已关闭标签: {name}")

    # ------------------------------------------------------------------ #
    # 全工作簿逐 Sheet 差异红点
    # ------------------------------------------------------------------ #
    def _refresh_bottom_bar_sheets(self) -> None:
        """两侧均加载后，重建底部标签栏并计算各 Sheet 差异状态。"""
        if not (self.left_wb and self.right_wb):
            return

        left_names = ExcelLoader.get_sheet_names(self.left_wb)
        right_names = ExcelLoader.get_sheet_names(self.right_wb)

        # 合并去重，保持左侧顺序优先，右侧独有的追加到末尾
        all_names: List[str] = list(left_names)
        for n in right_names:
            if n not in all_names:
                all_names.append(n)

        # 计算各 sheet 差异状态
        self._compute_sheet_diff_summary()

        diff_set = {n for n, has in self._sheet_diff_summary.items() if has}
        self.bottom_bar.set_sheets(all_names, diff_set)

        # 高亮当前 sheet
        current = self.left_sheet_name or self.right_sheet_name
        if current:
            self.bottom_bar.set_active(current)

    def _compute_sheet_diff_summary(self) -> None:
        """遍历所有 sheet，计算是否有差异（用于红点标记）。"""
        self._sheet_diff_summary = {}
        if not (self.left_wb and self.right_wb):
            return

        left_names = set(ExcelLoader.get_sheet_names(self.left_wb))
        right_names = set(ExcelLoader.get_sheet_names(self.right_wb))

        for name in left_names | right_names:
            if name not in left_names or name not in right_names:
                # 仅一侧存在 → 有差异
                self._sheet_diff_summary[name] = True
                continue
            try:
                lws = ExcelLoader.get_worksheet(self.left_wb, name)
                rws = ExcelLoader.get_worksheet(self.right_wb, name)
                ldata = ExcelLoader.extract_sheet_data(lws)
                rdata = ExcelLoader.extract_sheet_data(rws)
                diff = ExcelComparator.compare_sheets(
                    ldata, rdata, self.key_cols, self.ignore_cols
                )
                s = diff.stats
                self._sheet_diff_summary[name] = (
                    s["different"] + s["left_only"] + s["right_only"] > 0
                )
            except Exception:  # noqa: BLE001
                self._sheet_diff_summary[name] = True

    def _update_current_sheet_diff_status(self, result: DiffResult) -> None:
        """比较完成后，更新当前 sheet 的红点状态。"""
        current = self.left_sheet_name or self.right_sheet_name
        if not current:
            return
        s = result.stats
        has_diff = s["different"] + s["left_only"] + s["right_only"] > 0
        self._sheet_diff_summary[current] = has_diff
        self.bottom_bar.update_diff_status(
            {n for n, has in self._sheet_diff_summary.items() if has}
        )

    # ------------------------------------------------------------------ #
    # 工具方法
    # ------------------------------------------------------------------ #
    def update_status(self, text: str) -> None:
        """更新状态栏左侧提示文本。"""
        self.status_label.setText(text)

    def closeEvent(self, event) -> None:  # noqa: N802
        """关闭窗口：直接接受（Task 9 可能加入未保存确认）。"""
        event.accept()
