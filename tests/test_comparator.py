"""ExcelComparator 单元测试 — FR-02。

使用 unittest，多数用例直接构造 SheetData（worksheet 留空，比较器不使用该字段）；
合并单元格用例通过 ExcelLoader 在临时工作簿上端到端验证。
覆盖：默认顺序对齐、关键列对齐、left_only/right_only、行重排、忽略列、
公式字面量比较、不同列数、合并单元格、统计、差异单元格集合、空表保护。
"""

import os
import shutil
import tempfile
import unittest

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.core.comparator import DiffResult, ExcelComparator, RowPair
from src.core.excel_loader import ExcelLoader, SheetData


class ExcelComparatorTest(unittest.TestCase):
    """ExcelComparator 功能测试。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup)

    def _cleanup(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _path(self, name: str) -> str:
        return os.path.join(self.tmpdir, name)

    def _make_sheet(self, values, max_col=None) -> SheetData:
        """从二维字符串列表构造 SheetData。

        - 行自动补齐到 max_col 宽度，缺省单元格视为 ""。
        - worksheet 留空：比较器仅使用 max_row/max_col/values 字段。
        """
        if max_col is None:
            max_col = max((len(r) for r in values), default=0)
        padded = []
        for r in values:
            row = list(r)
            if len(row) < max_col:
                row = row + [""] * (max_col - len(row))
            padded.append(row)
        return SheetData(
            worksheet=None,
            max_row=len(padded),
            max_col=max_col,
            values=padded,
            merged_ranges=[],
            header_labels=[get_column_letter(i) for i in range(1, max_col + 1)],
        )

    # 1. 默认顺序对齐，完全相同 -> 全部 same
    def test_default_order_identical(self):
        left = self._make_sheet([["a", "b"], ["c", "d"]])
        right = self._make_sheet([["a", "b"], ["c", "d"]])
        result = ExcelComparator.compare_sheets(left, right)

        self.assertIsInstance(result, DiffResult)
        self.assertEqual(len(result.aligned_rows), 2)
        self.assertTrue(all(rp.status == "same" for rp in result.aligned_rows))
        self.assertEqual(result.stats, {"same": 2, "different": 0,
                                        "left_only": 0, "right_only": 0})
        self.assertEqual(result.diff_row_indices, [])
        self.assertEqual(result.diff_cell_set, set())
        self.assertEqual(result.max_col, 2)

    # 2. 默认顺序对齐，单个差异单元格 -> different，diff_cells 记录该列
    def test_default_order_one_diff_cell(self):
        left = self._make_sheet([["a", "b"], ["c", "d"]])
        right = self._make_sheet([["a", "b"], ["c", "X"]])
        result = ExcelComparator.compare_sheets(left, right)

        self.assertEqual(result.aligned_rows[0].status, "same")
        self.assertEqual(result.aligned_rows[1].status, "different")
        self.assertEqual(result.aligned_rows[1].diff_cells, [1])
        self.assertEqual(result.diff_row_indices, [1])
        self.assertEqual(result.diff_cell_set, {(1, 1)})

    # 3. 关键列对齐，左侧独占（右侧缺失某 key）
    def test_key_alignment_left_only(self):
        left = self._make_sheet([["1", "a"], ["2", "b"], ["3", "c"]])
        right = self._make_sheet([["1", "a"], ["3", "c"]])
        result = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        statuses = [rp.status for rp in result.aligned_rows]
        self.assertEqual(statuses, ["same", "left_only", "same"])
        self.assertEqual(result.aligned_rows[1].left_row, 1)
        self.assertIsNone(result.aligned_rows[1].right_row)
        self.assertEqual(result.aligned_rows[1].diff_cells, [])
        self.assertEqual(result.stats["left_only"], 1)

    # 4. 关键列对齐，右侧独占（左侧缺失某 key）
    def test_key_alignment_right_only(self):
        left = self._make_sheet([["1", "a"], ["3", "c"]])
        right = self._make_sheet([["1", "a"], ["2", "b"], ["3", "c"]])
        result = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # 左侧行先按 key 配对：left[0]<->right[0](key1)、left[1]<->right[2](key3)
        # 右侧未被消费的行（key2，位于 right 第 1 行）追加到末尾 -> right_only
        statuses = [rp.status for rp in result.aligned_rows]
        self.assertEqual(statuses, ["same", "same", "right_only"])
        self.assertEqual(result.aligned_rows[1].right_row, 2)  # key3 在右侧第 2 行
        self.assertEqual(result.aligned_rows[2].right_row, 1)  # key2 在右侧第 1 行
        self.assertIsNone(result.aligned_rows[2].left_row)
        self.assertEqual(result.stats["right_only"], 1)

    # 5. 关键列对齐，行被打乱重排 -> 按 key 匹配而非位置
    def test_key_alignment_reordered(self):
        left = self._make_sheet([["1", "a"], ["2", "b"], ["3", "c"]])
        right = self._make_sheet([["3", "c"], ["1", "a"], ["2", "b"]])
        result = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # 内容按 key 一致 -> 全部 same
        self.assertTrue(all(rp.status == "same" for rp in result.aligned_rows))
        self.assertEqual(result.stats["same"], 3)
        # 验证配对关系：left 0<->right 1, left 1<->right 2, left 2<->right 0
        pairs = {rp.left_row: rp.right_row for rp in result.aligned_rows}
        self.assertEqual(pairs, {0: 1, 1: 2, 2: 0})

    # 6. 忽略列：被忽略列的差异不导致 different
    def test_ignore_columns(self):
        left = self._make_sheet([["a", "b", "c"]])
        right = self._make_sheet([["a", "X", "c"]])
        result = ExcelComparator.compare_sheets(left, right, ignore_cols=[1])

        self.assertEqual(result.aligned_rows[0].status, "same")
        self.assertEqual(result.aligned_rows[0].diff_cells, [])
        self.assertEqual(result.diff_cell_set, set())
        # 不忽略时该列应判定为 different，作为对照
        result2 = ExcelComparator.compare_sheets(left, right)
        self.assertEqual(result2.aligned_rows[0].status, "different")
        self.assertEqual(result2.aligned_rows[0].diff_cells, [1])

    # 7. 公式按字面量比较："=A1" vs "=A2" -> different；相同公式 -> same
    def test_formula_literal_comparison(self):
        left = self._make_sheet([["=A1"]])
        right = self._make_sheet([["=A2"]])
        result = ExcelComparator.compare_sheets(left, right)

        self.assertEqual(result.aligned_rows[0].status, "different")
        self.assertEqual(result.aligned_rows[0].diff_cells, [0])
        self.assertEqual(result.diff_cell_set, {(0, 0)})

        # 相同公式字符串应判定 same
        left2 = self._make_sheet([["=A1"]])
        right2 = self._make_sheet([["=A1"]])
        result2 = ExcelComparator.compare_sheets(left2, right2)
        self.assertEqual(result2.aligned_rows[0].status, "same")

    # 8. 不同 max_col：缺省列视为 ""
    def test_different_max_col(self):
        left = self._make_sheet([["a", "b"]], max_col=2)
        right = self._make_sheet([["a", "b", "c"]], max_col=3)
        result = ExcelComparator.compare_sheets(left, right)

        self.assertEqual(result.max_col, 3)
        self.assertEqual(result.aligned_rows[0].status, "different")
        # 第 3 列（索引 2）：左侧缺省 "" vs 右侧 "c"
        self.assertIn(2, result.aligned_rows[0].diff_cells)
        self.assertEqual(result.diff_cell_set, {(0, 2)})

    # 9. 合并单元格值：SheetData.values 已填充合并区域，比较应正常工作
    def test_merged_cell_values(self):
        path = self._path("merged.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "M"
        ws.merge_cells("A1:B2")
        wb.save(path)

        wb2 = ExcelLoader.load_workbook(path)
        left = ExcelLoader.extract_sheet_data(wb2.active)
        # 合并区域 A1:B2 应被填充为 2x2 的 "M"
        self.assertEqual(left.max_row, 2)
        self.assertEqual(left.values, [["M", "M"], ["M", "M"]])

        # 用 left 的实际填充值构造等价右侧表（无合并区域）-> 全 same
        right = self._make_sheet(
            [list(r) for r in left.values], max_col=left.max_col
        )
        result = ExcelComparator.compare_sheets(left, right)
        self.assertTrue(all(rp.status == "same" for rp in result.aligned_rows))
        self.assertEqual(result.stats["same"], left.max_row)

    # 10. 统计正确性：same/different/left_only/right_only 计数匹配
    def test_stats_correctness(self):
        left = self._make_sheet([["1", "a"], ["2", "b"], ["3", "c"], ["4", "d"]])
        right = self._make_sheet([["1", "a"], ["2", "X"], ["5", "e"]])
        # key_cols=[0]：
        #   key1 双侧 a==a -> same
        #   key2 双侧 b!=X -> different
        #   key3 仅左 -> left_only
        #   key4 仅左 -> left_only
        #   key5 仅右 -> right_only
        result = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        self.assertEqual(result.stats, {"same": 1, "different": 1,
                                        "left_only": 2, "right_only": 1})
        # diff_row_indices 应包含所有非 same 行
        self.assertEqual(
            sorted(result.diff_row_indices), [1, 2, 3, 4]
        )

    # 11. diff_cell_set 坐标正确
    def test_diff_cell_set(self):
        left = self._make_sheet([["a", "b", "c"], ["x", "y", "z"]])
        right = self._make_sheet([["a", "X", "c"], ["x", "y", "Z"]])
        result = ExcelComparator.compare_sheets(left, right)

        # 第 0 行差异列 1；第 1 行差异列 2
        self.assertEqual(result.aligned_rows[0].diff_cells, [1])
        self.assertEqual(result.aligned_rows[1].diff_cells, [2])
        self.assertEqual(result.diff_cell_set, {(0, 1), (1, 2)})

    # 补充：空表保护（两侧 max_col=0、无行）
    def test_empty_sheets(self):
        left = self._make_sheet([])
        right = self._make_sheet([])
        result = ExcelComparator.compare_sheets(left, right)

        self.assertEqual(result.aligned_rows, [])
        self.assertEqual(result.diff_row_indices, [])
        self.assertEqual(result.diff_cell_set, set())
        self.assertEqual(result.stats, {"same": 0, "different": 0,
                                        "left_only": 0, "right_only": 0})
        self.assertEqual(result.max_col, 0)

    # 补充：默认顺序对齐，左右行数不等 -> 多余行分别 left_only/right_only
    def test_default_order_unequal_row_count(self):
        left = self._make_sheet([["a"], ["b"], ["c"]])
        right = self._make_sheet([["a"], ["b"]])
        result = ExcelComparator.compare_sheets(left, right)

        statuses = [rp.status for rp in result.aligned_rows]
        self.assertEqual(statuses, ["same", "same", "left_only"])
        self.assertEqual(result.stats["left_only"], 1)
        # 反向：右侧多出
        result2 = ExcelComparator.compare_sheets(right, left)
        self.assertEqual([rp.status for rp in result2.aligned_rows],
                         ["same", "same", "right_only"])

    # 补充：关键列对齐，两侧均有但内容不同 -> different 且 diff_cells 精确
    def test_key_alignment_different_content(self):
        left = self._make_sheet([["k1", "a", "b"], ["k2", "c", "d"]])
        right = self._make_sheet([["k2", "c", "X"], ["k1", "a", "b"]])
        result = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # k1 行 same，k2 行 different（列 2）
        by_key = {rp.left_row: rp for rp in result.aligned_rows}
        self.assertEqual(by_key[0].status, "same")
        self.assertEqual(by_key[1].status, "different")
        self.assertEqual(by_key[1].diff_cells, [2])
        self.assertEqual(by_key[1].right_row, 0)  # k2 在右侧位于第 0 行


if __name__ == "__main__":
    unittest.main()
