"""Excel 文件读写层 — ExcelLoader。

提供工作簿加载、工作表结构化数据提取、单元格样式/值复制与保存能力，
是后续比较器（Task 3）与合并器（Task 4）的基础设施。
对应需求：FR-01。
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
import zipfile
from copy import copy
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetData:
    """单个工作表的提取结果。

    字段说明：
    - worksheet: 原始 openpyxl Worksheet 对象，合并器需要它就地修改单元格/样式/合并区域
    - max_row / max_col: 数据区域的最大行/列数
    - values: 字符串化的单元格值二维表（0 索引），合并区域内所有单元格均填左上角值
      （data_only=False 加载，公式单元格保留公式字符串）
    - cached_values: 公式单元格的缓存计算值二维表（0 索引），由 data_only=True 副本提取。
      None 表示无缓存值（如合并后重新提取、或非公式场景），此时比较器仅比较公式。
    - merged_ranges: 合并单元格区域列表，元素为 (min_row, min_col, max_row, max_col)，1 索引
    - header_labels: 列标签，如 ["A","B","C",...]，长度等于 max_col
    """

    worksheet: Worksheet
    max_row: int
    max_col: int
    values: List[List[str]]
    cached_values: Optional[List[List[str]]] = None
    merged_ranges: List[Tuple[int, int, int, int]] = field(default_factory=list)
    header_labels: List[str] = field(default_factory=list)


class ExcelLoader:
    """Excel 读写工具类，全部为静态方法。"""

    @staticmethod
    def load_workbook(path: str) -> Workbook:
        """加载 .xlsx 工作簿。

        - data_only=False：保留公式字符串而非缓存计算值
        - keep_vba=False：不保留 VBA 宏
        - 注意：openpyxl 的 load_workbook 在当前版本没有 keep_styles 参数，
          样式默认就会被保留，因此此处无需也无法显式传入（PRD 中提到的
          keep_styles=True 并非真实参数，样式保留是默认行为）。
        - 仅支持 .xlsx；遇到 .xls 等旧格式抛出 ValueError。
        """
        if not str(path).lower().endswith(".xlsx"):
            raise ValueError(
                f"仅支持 .xlsx 格式文件，收到: {path}。"
                "如需支持旧版 .xls，请先用 Excel 另存为 .xlsx 后再处理。"
            )
        return load_workbook(path, data_only=False, keep_vba=False)

    @staticmethod
    def load_workbook_cached(path: str) -> Workbook:
        """加载 .xlsx 工作簿的 data_only 副本（用于提取公式缓存计算值）。

        data_only=True 时，公式单元格返回 Excel 上次保存时缓存的计算值
        （而非公式字符串）。配合 load_workbook 使用以同时获取公式与计算值。
        注意：若文件从未被 Excel 打开保存，缓存值可能为 None。
        """
        if not str(path).lower().endswith(".xlsx"):
            raise ValueError(
                f"仅支持 .xlsx 格式文件，收到: {path}。"
                "如需支持旧版 .xls，请先用 Excel 另存为 .xlsx 后再处理。"
            )
        return load_workbook(path, data_only=True, keep_vba=False)

    @staticmethod
    def get_sheet_names(wb: Workbook) -> List[str]:
        """返回工作簿中所有工作表名称。"""
        return wb.sheetnames

    @staticmethod
    def get_worksheet(wb: Workbook, name: str) -> Worksheet:
        """按名称获取工作表。"""
        return wb[name]

    @staticmethod
    def get_merged_ranges(ws: Worksheet) -> List[Tuple[int, int, int, int]]:
        """返回工作表的合并单元格区域列表，元素为 (min_row, min_col, max_row, max_col)。"""
        ranges: List[Tuple[int, int, int, int]] = []
        for mr in ws.merged_cells.ranges:
            ranges.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
        return ranges

    @staticmethod
    def extract_sheet_data(
        ws: Worksheet, cached_ws: Optional[Worksheet] = None
    ) -> SheetData:
        """从 Worksheet 提取结构化数据 SheetData。

        - ws: data_only=False 的工作表，公式单元格保留公式字符串。
        - cached_ws: data_only=True 的同名工作表，用于提取公式单元格的缓存计算值。
          传入后 cached_values 字段被填充；为 None 时 cached_values=None
          （比较器退化为仅比较公式）。
        """
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        merged_ranges = ExcelLoader.get_merged_ranges(ws)

        # 空表保护：行或列为 0 时返回空结构，避免后续循环越界
        if max_row <= 0 or max_col <= 0:
            return SheetData(
                worksheet=ws,
                max_row=max_row,
                max_col=max_col,
                values=[],
                cached_values=None,
                merged_ranges=merged_ranges,
                header_labels=[],
            )

        # 先按单元格原值构建字符串二维表
        values: List[List[str]] = [
            ["" for _ in range(max_col)] for _ in range(max_row)
        ]
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                values[r - 1][c - 1] = ExcelLoader._stringify(ws.cell(r, c).value)

        # 合并单元格：用左上角值填充区域内所有单元格，保证比较时一致
        for mr in ws.merged_cells.ranges:
            min_r, min_c = mr.min_row, mr.min_col
            max_r, max_c = mr.max_row, mr.max_col
            top_value = ExcelLoader._stringify(ws.cell(min_r, min_c).value)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    # 越界保护：合并区域理论上不会超出 max_row/max_col，此处防御性裁剪
                    if 1 <= r <= max_row and 1 <= c <= max_col:
                        values[r - 1][c - 1] = top_value

        # 缓存计算值（data_only=True 副本）：用于检测"公式相同但计算值不同"
        cached_values: Optional[List[List[str]]] = None
        if cached_ws is not None:
            cached_values = [
                ["" for _ in range(max_col)] for _ in range(max_row)
            ]
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cached_values[r - 1][c - 1] = ExcelLoader._stringify(
                        cached_ws.cell(r, c).value
                    )
            # 合并单元格同样用左上角缓存值填充
            for mr in cached_ws.merged_cells.ranges:
                min_r, min_c = mr.min_row, mr.min_col
                max_r, max_c = mr.max_row, mr.max_col
                top_value = ExcelLoader._stringify(
                    cached_ws.cell(min_r, min_c).value
                )
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        if 1 <= r <= max_row and 1 <= c <= max_col:
                            cached_values[r - 1][c - 1] = top_value

        header_labels = [get_column_letter(i) for i in range(1, max_col + 1)]

        return SheetData(
            worksheet=ws,
            max_row=max_row,
            max_col=max_col,
            values=values,
            cached_values=cached_values,
            merged_ranges=merged_ranges,
            header_labels=header_labels,
        )

    @staticmethod
    def _stringify(value) -> str:
        """将单元格值统一转为字符串。

        - None -> ""
        - bool -> "TRUE"/"FALSE"（bool 是 int 子类，必须先于 int 判断）
        - 公式（以 "=" 开头的字符串）原样保留
        - ArrayFormula / DataTableFormula 等公式对象：取其 text 属性
          （数组公式文本，如 "=SUM(A1:A2)"），避免 str() 输出对象地址
        - 其它类型直接 str()
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        # ArrayFormula 等公式对象：取 text 属性，避免 str() 输出对象地址
        text = getattr(value, "text", None)
        if text is not None:
            return str(text)
        return str(value)

    @staticmethod
    def copy_cell_value(
        src_cell, tgt_cell, cached_value: Optional[str] = None
    ) -> None:
        """复制单元格值（公式 + 缓存值）。

        - 普通公式（以 "=" 开头的字符串）原样写入，openpyxl 自动识别为公式。
        - 数组公式（ArrayFormula 对象）：提取其 text（公式文本）以普通公式
          字符串写入目标单元格，避免直接赋值对象导致 ref 仍指向源位置、
          目标单元格引用错误。
        - 其它值（含 None）直接写入。
        - cached_value：源单元格的缓存计算值（来自 data_only=True 副本）。
          写入工作簿的 _pending_cached_values，按 {sheet_name: {coord: value}} 结构存储，
          保存时注入 XML 的 <v> 标签。
        """
        value = src_cell.value
        # ArrayFormula 等公式对象：取 text 属性以普通公式字符串写入
        text = getattr(value, "text", None)
        if text is not None:
            tgt_cell.value = text
        else:
            tgt_cell.value = value

        # 记录缓存值，保存时注入 XML
        if cached_value is not None:
            wb = tgt_cell.parent.parent if tgt_cell.parent else None
            ws_name = tgt_cell.parent.title if tgt_cell.parent else None
            if wb is not None and ws_name:
                if not hasattr(wb, "_pending_cached_values"):
                    wb._pending_cached_values = {}  # type: ignore[attr-defined]
                if ws_name not in wb._pending_cached_values:
                    wb._pending_cached_values[ws_name] = {}  # type: ignore[attr-defined]
                wb._pending_cached_values[ws_name][tgt_cell.coordinate] = cached_value  # type: ignore[attr-defined]

    @staticmethod
    def save_workbook(wb: Workbook, path: str) -> None:
        """保存工作簿到指定路径，并注入缓存计算值到公式单元格。

        - 先用 openpyxl 保存到临时文件
        - 解析 workbook.xml 建立 sheet_name -> sheet_xml_path 映射
        - 扫描每个 sheet XML，为有缓存值的公式单元格替换/插入 <v> 标签
        - 移动到目标路径
        - fullCalcOnLoad 设为 False（保留缓存值，不强制重算）
        """
        try:
            wb.calculation.fullCalcOnLoad = False
        except Exception:  # noqa: BLE001
            pass

        cached_values: Dict[str, Dict[str, str]] = getattr(
            wb, "_pending_cached_values", {}
        )

        if not cached_values:
            wb.save(path)
            return

        # 保存到临时文件，再注入缓存值后写到目标路径
        tmp_fd, tmp_path = tempfile.mkstemp(
            suffix=".xlsx", prefix="excelmerge_"
        )
        os.close(tmp_fd)
        try:
            wb.save(tmp_path)
            ExcelLoader._inject_cached_values(tmp_path, path, cached_values)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    @staticmethod
    def _get_sheet_xml_mapping(
        zin: zipfile.ZipFile,
    ) -> Dict[str, str]:
        """解析 workbook.xml 和 workbook.xml.rels，返回 {sheet_name: xml_path} 映射。

        xlsx 内部结构:
        - xl/workbook.xml 的 <sheets><sheet name="..." r:id="rIdX"/></sheets>
        - xl/_rels/workbook.xml.rels 的 <Relationship Id="rIdX" Target="worksheets/sheetN.xml"/>
        """
        sheet_name_to_rid: Dict[str, str] = {}
        rid_to_target: Dict[str, str] = {}

        # 解析 workbook.xml 获取 sheet_name -> rId
        try:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            # 匹配所有 <sheet ...> 或 <sheet .../> 标签
            for m in re.finditer(r'<sheet\s[^>]*?/?>', wb_xml):
                tag = m.group(0)
                name_match = re.search(r'name="([^"]+)"', tag)
                rid_match = re.search(r'r:id="([^"]+)"', tag)
                if name_match and rid_match:
                    sheet_name_to_rid[name_match.group(1)] = rid_match.group(1)
        except Exception:
            return {}

        # 解析 workbook.xml.rels 获取 rId -> Target
        try:
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            # 匹配所有 <Relationship ...> 或 <Relationship .../> 标签
            for m in re.finditer(r'<Relationship\s[^>]*?/?>', rels_xml):
                tag = m.group(0)
                id_match = re.search(r'Id="([^"]+)"', tag)
                target_match = re.search(r'Target="([^"]+)"', tag)
                if not id_match or not target_match:
                    continue
                rid = id_match.group(1)
                target = target_match.group(1)
                # Target 可能是 "/xl/worksheets/sheet1.xml"（绝对路径）或
                # "worksheets/sheet1.xml"（相对于 xl/ 的路径）
                # 统一去掉前导 / 前缀
                if target.startswith("/"):
                    target = target[1:]
                # 如果不是以 xl/ 开头，加上 xl/ 前缀（相对于 xl/ 的路径）
                if not target.startswith("xl/"):
                    target = "xl/" + target
                rid_to_target[rid] = target
        except Exception:
            return {}

        # 组装 sheet_name -> xml_path
        mapping: Dict[str, str] = {}
        for name, rid in sheet_name_to_rid.items():
            if rid in rid_to_target:
                mapping[name] = rid_to_target[rid]
        return mapping

    @staticmethod
    def _inject_cached_values(
        src_path: str, dst_path: str, cached_values: Dict[str, Dict[str, str]]
    ) -> None:
        """注入缓存计算值到 xlsx 文件的公式单元格 XML。

        - cached_values 格式: {sheet_name: {coordinate: cached_value}}
        - 解析 sheet 名与 XML 路径映射后，对每个 sheet 分别注入对应缓存值。
        - 若无法解析映射（异常情况），退化为兼容旧逻辑：对所有 sheet XML 用全部缓存值注入
          （单 sheet 文件可正常工作）。
        """
        with zipfile.ZipFile(src_path, "r") as zin:
            sheet_mapping = ExcelLoader._get_sheet_xml_mapping(zin)

            with zipfile.ZipFile(dst_path, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    is_sheet_xml = (
                        item.filename.startswith("xl/worksheets/sheet")
                        and item.filename.endswith(".xml")
                    )
                    if is_sheet_xml:
                        xml = data.decode("utf-8")
                        # 找到该 XML 对应的 sheet 名
                        target_sheet = None
                        for name, xml_path in sheet_mapping.items():
                            if xml_path == item.filename:
                                target_sheet = name
                                break
                        if target_sheet and target_sheet in cached_values:
                            value_map = cached_values[target_sheet]
                            xml = ExcelLoader._inject_xml_cached_values(
                                xml, value_map
                            )
                        elif not sheet_mapping:
                            # 无法解析映射时的兼容回退：单sheet场景下把所有缓存值都尝试注入
                            for vmap in cached_values.values():
                                xml = ExcelLoader._inject_xml_cached_values(
                                    xml, vmap
                                )
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)

    @staticmethod
    def _inject_xml_cached_values(
        xml: str, value_map: Dict[str, str]
    ) -> str:
        """注入缓存值到单个 sheet XML。

        匹配公式单元格 <c ...><f...>...</f>...<v.../></c> 或 <c ...><f...>...</f>...<v>...</v></c>，
        将 <v> 替换为对应缓存值；若无 <v> 则插入。正确处理 openpyxl 生成的自闭合 <v />。
        """
        # 匹配c标签，支持 r 属性不在第一个位置，支持自闭合<v/>
        pattern = re.compile(
            r'<c\s(?P<attrs>[^>]*\br="(?P<ref>[A-Z]+\d+)"[^>]*)>'
            r'(?P<inner>.*?)</c>',
            re.DOTALL
        )

        def replacer(m: re.Match) -> str:
            ref = m.group("ref")
            if ref not in value_map:
                return m.group(0)
            attrs = m.group("attrs")
            inner = m.group("inner")
            cached = value_map[ref]

            # 必须是公式单元格（包含 <f 标签）
            if "<f" not in inner:
                return m.group(0)

            # 移除所有现有的 v 标签（包括自闭合 <v/> 和成对 <v>...</v>）
            inner = re.sub(r'<v\s[^>]*/>', '', inner)
            inner = re.sub(r'<v[^>]*/>', '', inner)
            inner = re.sub(r'<v[^>]*>[^<]*</v>', '', inner)

            # 判断是否为字符串结果
            is_str = not (
                cached.lstrip("-").replace(".", "", 1).isdigit()
            )

            # 处理 <c> 标签上的 t 属性
            if is_str:
                if 't="' in attrs:
                    attrs = re.sub(r't="[^"]*"', 't="str"', attrs)
                else:
                    attrs = attrs + ' t="str"'
            else:
                attrs = re.sub(r'\s*t="[^"]*"', '', attrs)

            # 在 </f> 后面插入新的 <v> 标签
            new_v = f"<v>{cached}</v>"
            # 找到 </f> 的位置，在其后插入
            f_close_match = re.search(r'</f\s*>', inner)
            if f_close_match:
                insert_pos = f_close_match.end()
                inner = inner[:insert_pos] + new_v + inner[insert_pos:]
            else:
                # 找不到 </f>，追加到 inner 末尾（不应该发生，但防御性处理）
                inner = inner + new_v

            return f"<c {attrs}>{inner}</c>"

        return pattern.sub(replacer, xml)

    @staticmethod
    def copy_cell_style(src_cell, tgt_cell) -> None:
        """深拷贝源单元格的样式到目标单元格。

        复制项：font、fill、border、alignment、number_format。
        使用 copy.copy 生成独立样式对象，避免目标修改连锁影响源单元格。
        """
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.alignment = copy(src_cell.alignment)
        tgt_cell.number_format = src_cell.number_format
