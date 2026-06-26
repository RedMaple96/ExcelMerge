"""ExcelMerger 单元测试 — FR-04。

使用 unittest，测试数据通过 openpyxl 在临时目录中动态生成 .xlsx 文件，
并通过 ExcelLoader.extract_sheet_data 与 ExcelComparator.compare_sheets
构造真实 DiffResult，端到端验证各合并策略。
覆盖：右覆盖(different/right_only)、左覆盖、append_rows、公式存活、
样式复制、合并单元格重建、逐项手动合并、保存后重读。
"""

import os
import shutil
import tempfile
import unittest

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.core.comparator import ExcelComparator
from src.core.excel_loader import ExcelLoader, SheetData
from src.core.merger import ExcelMerger


def _merged_tuples(ws):
    """把 worksheet.merged_cells.ranges 转为 (min_row,min_col,max_row,max_col) 元组列表。"""
    return [(mr.min_row, mr.min_col, mr.max_row, mr.max_col) for mr in ws.merged_cells.ranges]


class ExcelMergerTest(unittest.TestCase):
    """ExcelMerger 功能测试。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup)

    def _cleanup(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _path(self, name: str) -> str:
        return os.path.join(self.tmpdir, name)

    # 1. 右覆盖 - different 行：左 A1=1, 右 A1=2 -> 右覆盖后左 A1==2 且样式与右一致
    def test_merge_right_to_left_different_row(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 2
        ws_r["A1"].font = Font(bold=True)
        ws_r["A1"].fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right)

        # 比较结果应为 different
        self.assertEqual(diff.aligned_rows[0].status, "different")
        self.assertEqual(diff.aligned_rows[0].diff_cells, [0])

        ExcelMerger.merge_right_to_left(diff, left, right)

        self.assertEqual(left.worksheet["A1"].value, 2)
        # 样式应与右一致
        self.assertTrue(left.worksheet["A1"].font.bold)
        self.assertEqual(
            left.worksheet["A1"].fill.fgColor.rgb, right.worksheet["A1"].fill.fgColor.rgb
        )

    # 2. 右覆盖 - right_only 追加：左 1 行，右 2 行(key 对齐后第 2 行 right_only)
    def test_merge_right_to_left_right_only_append(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "k1"
        ws_l["B1"] = "a"
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "k1"
        ws_r["B1"] = "a"
        ws_r["A2"] = "k2"
        ws_r["B2"] = "b"
        ws_r["B2"].font = Font(italic=True)

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # 第 0 行 same，第 1 行 right_only
        self.assertEqual(diff.aligned_rows[0].status, "same")
        self.assertEqual(diff.aligned_rows[1].status, "right_only")

        ExcelMerger.merge_right_to_left(diff, left, right)

        # 左侧应追加到 2 行
        self.assertEqual(left.worksheet.max_row, 2)
        self.assertEqual(left.worksheet["A2"].value, "k2")
        self.assertEqual(left.worksheet["B2"].value, "b")
        # 样式应来自右
        self.assertTrue(left.worksheet["B2"].font.italic)

    # 3. 左覆盖：镜像，左覆盖后右 == 左
    def test_merge_left_to_right(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 5
        ws_l["A1"].font = Font(bold=True)
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 9

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right)
        self.assertEqual(diff.aligned_rows[0].status, "different")

        ExcelMerger.merge_left_to_right(diff, left, right)

        # 右 A1 应被左侧覆盖
        self.assertEqual(right.worksheet["A1"].value, 5)
        self.assertTrue(right.worksheet["A1"].font.bold)

    # 3b. 左覆盖 - left_only 追加：左 2 行，右 1 行 -> 左第 2 行追加到右
    def test_merge_left_to_right_left_only_append(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "k1"
        ws_l["A2"] = "k2"
        ws_l["B2"] = "extra"
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "k1"

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right, key_cols=[0])

        # 第 0 行 same，第 1 行 left_only
        self.assertEqual(diff.aligned_rows[0].status, "same")
        self.assertEqual(diff.aligned_rows[1].status, "left_only")

        ExcelMerger.merge_left_to_right(diff, left, right)

        self.assertEqual(right.worksheet.max_row, 2)
        self.assertEqual(right.worksheet["A2"].value, "k2")
        self.assertEqual(right.worksheet["B2"].value, "extra")

    # 4. append_rows：仅追加 right_only，不动 different 行
    def test_append_rows_only_right_only(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1  # 与右不同 -> different
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 2  # different
        ws_r["A2"] = "k2"  # right_only

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right)

        # 第 0 行 different，第 1 行 right_only
        self.assertEqual(diff.aligned_rows[0].status, "different")
        self.assertEqual(diff.aligned_rows[1].status, "right_only")

        ExcelMerger.append_rows(diff, left, right)

        # different 行左侧原值应保持不变
        self.assertEqual(left.worksheet["A1"].value, 1)
        # right_only 行应追加到左侧
        self.assertEqual(left.worksheet.max_row, 2)
        self.assertEqual(left.worksheet["A2"].value, "k2")

    # 5. 公式存活：右 C1="=A1+B1"，右覆盖后左 C1 == "=A1+B1"（字符串）
    def test_formula_preserved_after_merge(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        ws_l["B1"] = 2
        ws_l["C1"] = "old"
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 1
        ws_r["B1"] = 2
        ws_r["C1"] = "=A1+B1"

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right)

        # 仅 C1 不同
        self.assertEqual(diff.aligned_rows[0].status, "different")
        self.assertEqual(diff.aligned_rows[0].diff_cells, [2])

        ExcelMerger.merge_right_to_left(diff, left, right)

        # 公式应原样写入，而非被计算值替代
        self.assertEqual(left.worksheet["C1"].value, "=A1+B1")

    # 6. 样式复制：右单元格红色填充+加粗，右覆盖后左单元格 fill==右 fill、font.bold==True
    def test_style_copied_after_merge(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "x"
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "y"
        ws_r["A1"].fill = PatternFill(
            start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
        )
        ws_r["A1"].font = Font(bold=True)

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        diff = ExcelComparator.compare_sheets(left, right)

        ExcelMerger.merge_right_to_left(diff, left, right)

        self.assertEqual(left.worksheet["A1"].value, "y")
        self.assertTrue(left.worksheet["A1"].font.bold)
        self.assertEqual(
            left.worksheet["A1"].fill.fgColor.rgb, right.worksheet["A1"].fill.fgColor.rgb
        )

    # 7. 合并单元格重建：右表某行 A:B 合并，追加该行后左表对应行也有相同合并
    def test_merged_cell_rebuilt_on_append(self):
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = "x"  # 与右第 1 行不同 -> different
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = "y"  # different(与左配对)
        ws_r["A2"] = "M"  # right_only，且 A2:B2 合并
        ws_r.merge_cells("A2:B2")

        left = ExcelLoader.extract_sheet_data(ws_l)
        right = ExcelLoader.extract_sheet_data(ws_r)
        # 右 A2:B2 合并后 values[1] = ["M","M"]；左只有 1 行 -> 第 1 行 right_only
        diff = ExcelComparator.compare_sheets(left, right)

        self.assertEqual(diff.aligned_rows[0].status, "different")
        self.assertEqual(diff.aligned_rows[1].status, "right_only")

        ExcelMerger.merge_right_to_left(diff, left, right)

        # 追加行应在左侧第 2 行，且 A2:B2 合并
        self.assertEqual(left.worksheet.max_row, 2)
        self.assertEqual(left.worksheet["A2"].value, "M")
        self.assertIn((2, 1, 2, 2), _merged_tuples(left.worksheet))

    # 8. copy_row_to_other / copy_single_cell：手动复制单行/单格
    def test_copy_row_to_other_and_single_cell(self):
        wb_s = Workbook()
        ws_s = wb_s.active
        ws_s["A1"] = "s1"
        ws_s["B1"] = "s2"
        ws_s["A1"].font = Font(bold=True)
        ws_s["B1"].fill = PatternFill(
            start_color="FF00FF00", end_color="FF00FF00", fill_type="solid"
        )
        wb_t = Workbook()
        ws_t = wb_t.active
        ws_t["A1"] = None
        ws_t["B1"] = None

        source = ExcelLoader.extract_sheet_data(ws_s)
        target = ExcelLoader.extract_sheet_data(ws_t)

        # 复制整行
        ExcelMerger.copy_row_to_other(source, 0, target, 0, 2)
        self.assertEqual(target.worksheet["A1"].value, "s1")
        self.assertEqual(target.worksheet["B1"].value, "s2")
        self.assertTrue(target.worksheet["A1"].font.bold)
        self.assertIn("00FF00", target.worksheet["B1"].fill.fgColor.rgb)

        # 复制单个单元格：源 A2 -> 目标 A2
        ws_s["A2"] = "single"
        ws_s["A2"].font = Font(italic=True)
        # 重新提取以包含新行
        source = ExcelLoader.extract_sheet_data(ws_s)
        ws_t["A2"] = None
        target = ExcelLoader.extract_sheet_data(ws_t)
        ExcelMerger.copy_single_cell(source, 1, 0, target, 1)
        self.assertEqual(target.worksheet["A2"].value, "single")
        self.assertTrue(target.worksheet["A2"].font.italic)

    # 9. 保存后重读：merge + save + reload，验证修改持久且文件未损坏
    def test_save_and_reload_after_merge(self):
        left_path = self._path("left.xlsx")
        right_path = self._path("right.xlsx")
        wb_l = Workbook()
        ws_l = wb_l.active
        ws_l["A1"] = 1
        ws_l["B1"] = "keep"
        wb_l.save(left_path)
        wb_r = Workbook()
        ws_r = wb_r.active
        ws_r["A1"] = 99
        ws_r["B1"] = "keep"
        ws_r["A2"] = "new_row"
        wb_r.save(right_path)

        wb_l2 = ExcelLoader.load_workbook(left_path)
        wb_r2 = ExcelLoader.load_workbook(right_path)
        left = ExcelLoader.extract_sheet_data(wb_l2.active)
        right = ExcelLoader.extract_sheet_data(wb_r2.active)
        diff = ExcelComparator.compare_sheets(left, right)

        ExcelMerger.merge_right_to_left(diff, left, right)
        ExcelLoader.save_workbook(wb_l2, left_path)

        # 重新加载，验证修改持久
        wb_l3 = ExcelLoader.load_workbook(left_path)
        ws_l3 = wb_l3.active
        self.assertEqual(ws_l3["A1"].value, 99)  # 右覆盖生效
        self.assertEqual(ws_l3["B1"].value, "keep")  # 未变的列保持
        self.assertEqual(ws_l3["A2"].value, "new_row")  # 追加行持久
        self.assertEqual(ws_l3.max_row, 2)

    # 补充：_find_merged_ranges_for_row 仅返回以该行为左上角的区域
    def test_find_merged_ranges_for_row(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")  # min_row=1
        ws.merge_cells("A3:B4")  # min_row=3，跨 2 行
        data = ExcelLoader.extract_sheet_data(ws)

        self.assertEqual(ExcelMerger._find_merged_ranges_for_row(data, 1), [(1, 1, 1, 2)])
        self.assertEqual(ExcelMerger._find_merged_ranges_for_row(data, 3), [(3, 1, 4, 2)])
        # 第 2、4 行不是任何合并区域左上角
        self.assertEqual(ExcelMerger._find_merged_ranges_for_row(data, 2), [])
        self.assertEqual(ExcelMerger._find_merged_ranges_for_row(data, 4), [])

    # 补充：_append_row 返回追加行号(1-based)，且跨行合并区域按行跨度重建
    def test_append_row_returns_row_and_rebuilds_span(self):
        wb_s = Workbook()
        ws_s = wb_s.active
        ws_s["A1"] = "top"
        ws_s.merge_cells("A1:B2")  # 跨 2 行的合并，min_row=1
        wb_t = Workbook()
        ws_t = wb_t.active
        ws_t["A1"] = "exists"  # target 已有 1 行

        source = ExcelLoader.extract_sheet_data(ws_s)
        target = ExcelLoader.extract_sheet_data(ws_t)

        append_row = ExcelMerger._append_row(source, 0, target, 2)
        # target 原 max_row=1，追加位置应为 2
        self.assertEqual(append_row, 2)
        # 跨行合并应保持行跨度：原 A1:B2(跨度1) -> A2:B3
        self.assertIn((2, 1, 3, 2), _merged_tuples(target.worksheet))
        self.assertEqual(target.worksheet["A2"].value, "top")


if __name__ == "__main__":
    unittest.main()
